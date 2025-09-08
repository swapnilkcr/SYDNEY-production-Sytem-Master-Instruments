import base64
import bcrypt
import csv
import openpyxl
import json
import sqlite3
from datetime import datetime
import pandas as pd
import io
import os
from http.server import BaseHTTPRequestHandler, HTTPServer
from datetime import datetime
import asyncio
import websockets
from urllib.parse import urlparse, parse_qs
import urllib.parse
import time
from config import ENV, PORT,BASE_URL  # Import environment variables
import gzip
import json
from io import BytesIO  
import sqlite3
from datetime import datetime
import json
from config import DB_NAME
from datetime import timedelta
from jinja2 import Environment, FileSystemLoader

# Jinja2 environment: looks inside "templates" folder
env = Environment(loader=FileSystemLoader("templates"))
print("DB being used:", os.path.abspath(DB_NAME))




print(f"🚀 Environment: {ENV}")
print(f"🌐 Running on port: {PORT}")


PRODUCT_PHOTOS = 'PRODUCT PHOTOS'  # Folder to store uploaded PDFs
os.makedirs(PRODUCT_PHOTOS, exist_ok=True)  # Ensure the folder exists

# ✅ Enable WAL mode for better concurrency (run once at startup)
try:
    conn = sqlite3.connect(DB_NAME, timeout=30)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.close()
    print("✅ SQLite WAL mode enabled")
except Exception as e:
    print("⚠️ Could not enable WAL mode:", e)

# ✅ Helper function to always return a safe connection
def get_db_connection():
    """Return a SQLite connection with timeout & autocommit enabled."""
    return sqlite3.connect(DB_NAME, timeout=30, isolation_level=None)






def execute_with_retry(cursor, query, params=(), max_retries=5, delay=1):
    for attempt in range(max_retries):
        try:
            cursor.execute(query, params)
            return
        except sqlite3.OperationalError as e:
            if "database is locked" in str(e).lower():
                print(f"Database is locked. Retrying ({attempt + 1}/{max_retries})...")
                time.sleep(delay)
            else:
                raise
    raise sqlite3.OperationalError("Database is locked after multiple retries")

def commit_with_retry(conn, max_retries=5, delay=1):
    for attempt in range(max_retries):
        try:
            conn.commit()
            return
        except sqlite3.OperationalError as e:
            if "database is locked" in str(e).lower():
                print(f"Commit failed. Retrying ({attempt + 1}/{max_retries})...")
                time.sleep(delay)
            else:
                raise
    raise sqlite3.OperationalError("Commit failed after multiple retries")


#DB_NAME = 'clock_in_management.db'



'''async def send_update(action, staff_name, job_id, timestamp):
    """Send a WebSocket update when a job starts/stops."""
    message = {
        "action": action,
        "staffName": staff_name,
        "jobId": job_id,
        "timestamp": timestamp
    }
    try:
        async with websockets.connect("ws://127.0.0.1:8765") as websocket:
            await websocket.send(json.dumps(message))
    except Exception as e:
        print(f"Error sending WebSocket update: {e}")'''





# Format timestamp to 'YYYY-MM-DD HH:MM:SS'
def get_current_timestamp():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


# Utility to ensure file exists
def ensure_file_exists(file_path, header=None):
    if not os.path.isfile(file_path):
        with open(file_path, 'w') as f:
            if header:
                f.write(header + '\n')

# Ensure staff.csv and jobs.csv exist
ensure_file_exists('staff.csv', header='StaffName')  # Add 'StaffName' as header for staff file


    

# Load staff names from a CSV file
def load_staff_from_csv(file_path):
    staff = []
    try:
        with open(file_path, mode='r') as csvfile:
            print(f"Attempting to read file: {file_path}")
            reader = csv.reader(csvfile)
            # Skip the header row if the CSV has one
            next(reader, None)
            for row in reader:
                print(f"Row read: {row}")
                if row:  # Ensure the row is not empty
                    staff.append(row[0])  # Assuming the first column contains staff names
        print(f"Staff loaded successfully: {staff}") 
    except Exception as e:
        print(f"Error reading staff CSV file: {e}")
    return staff



# Calculate Estimated Time from Av and QTY
def calculate_estimated_time(job_id):
    """Calculate the estimated time for a job based on AV and QTY columns from the database."""
    try:
        print(f"DEBUG: Called calculate_estimated_time() with job_id={job_id}")  # ✅ Ensure function is called

        conn = get_db_connection()

        print("Using DB at:", os.path.abspath(DB_NAME))
        cursor = conn.cursor()

        # Fetch AV and QTY for the given job_id
        cursor.execute('SELECT AV, QTY FROM PN_DATA WHERE PN = ?', (job_id,))
        result = cursor.fetchone()
        conn.close()

        if result:
            av = float(result[0]) if result[0] is not None else 0.0
            qty = float(result[1]) if result[1] is not None else 0.0
            estimated_time = round(av * qty, 2)

            # ✅ Print debug info
            print(f"DEBUG: JobID={job_id}, AV={av}, QTY={qty}, Estimated Time={estimated_time}")

            return estimated_time
        else:
            print(f"DEBUG: No matching record found for JobID={job_id}")
            return 0.0  # Ensure function always returns a value

    except Exception as e:
        print(f"Error fetching estimated time: {e}")
        return 0.0

       
# Calculate Total Hours Worked by all users for a job
def get_total_hours_worked(job_id):
    conn = get_db_connection()

    cursor = conn.cursor()
    cursor.execute('''
        SELECT COALESCE(SUM(
            CASE WHEN StopTime IS NOT NULL THEN 
                (strftime('%s', StopTime) - strftime('%s', StartTime)) / 3600.0
            ELSE
                (strftime('%s', 'now') - strftime('%s', StartTime)) / 3600.0  -- Ongoing job
            END
        ), 0.0)
        FROM ClockInOut WHERE JobID = ? AND StaffName != 'QC'
    ''', (job_id,))
    result = cursor.fetchone()
    conn.close()
    #print(f"Debug: SQL Query Result for JobID {job_id}: {result}")  # Debugging line
    return float(result[0]) if result and result[0] is not None else 0.0


def get_job_details(job_id):
    """Fetch customer name, drawing number, cell number, quantity, and required date from the database based on job_id (PN)."""
    try:
        # Connect to the database
        conn = get_db_connection()

        cursor = conn.cursor()

        # Query to fetch all required job details
        cursor.execute('''
            SELECT CUST, "DRAW NO", "NO/CELL", QTY, "REQU-DATE"
            FROM PN_DATA
            WHERE PN = ?
        ''', (job_id,))
        result = cursor.fetchone()

        # Close the database connection
        conn.close()

        # Return a dictionary with fetched values or "Unknown" if not found
        return {
            "customerName": result[0] if result and result[0] else " No data",
            "drawingNumber": result[1] if result and result[1] else " ",
            "cellNo": result[2] if result and result[2] else " ",
            "quantity": result[3] if result and result[3] else " ",
            "requiredDate": result[4] if result and result[4] else " "
        }
    except Exception as e:
        print(f"Error fetching job details from the database: {e}")
        return {
            "customerName": "Unknown",
            "drawingNumber": "Unknown",
            "cellNo": "Unknown",
            "quantity": "Unknown",
            "requiredDate": "Unknown"
        }

    

def get_av_by_stock_code(stock_code):
    """Fetch AV and DrawNo based on STOCKCODE from the database."""
    try:
        conn = get_db_connection()

        cursor = conn.cursor()

        cursor.execute('SELECT AV, DrawNo FROM MergedData WHERE STOCKCODE = ?', (stock_code,))
        result = cursor.fetchone()
        conn.close()

        return {'avValue': result[0], 'drawNo': result[1]} if result else None
    except Exception as e:
        print(f"Error fetching AV from STOCKCODE: {e}")
        return None
    
def get_job_work_details(job_id):
    try:
        print(f"DEBUG: Called get_job_work_details() with job_id={job_id}")

        conn = get_db_connection()

        cursor = conn.cursor()

        # Fetch estimated time
        estimated_time = calculate_estimated_time(job_id)
        print(f"DEBUG: Estimated time = {estimated_time}")

        # Fetch summary of job work from jobsfinished or ClockInOut or wherever
        cursor.execute("""
            SELECT remainingTime, totalHoursWorked, totalLaborCost, status
            FROM JOBSFINISHED
            WHERE JobID = ?
            LIMIT 1
        """, (job_id,))
        row = cursor.fetchone()

        if not row:
            print(f"DEBUG: No matching record found for JobID={job_id}")
            return {
                'jobId': job_id,
                'estimatedTime': float(estimated_time or 0),
                'users': [],
                'remainingTime': 0,
                'totalHoursWorked': 0,
                'totalLaborCost': 0,
                'status': 'Not Found'
            }

        # Unpack safely
        remaining_time = float(row[0] or 0)
        total_worked = float(row[1] or 0)
        total_labor_cost = float(row[2] or 0)
        status = row[3] or "Unknown"

        # Fetch staff time breakdown from ClockInOut
        cursor.execute("""
            SELECT StaffName, 
                   ROUND(SUM((julianday(COALESCE(StopTime, CURRENT_TIMESTAMP)) - julianday(StartTime)) * 24.0), 2)
            FROM ClockInOut
            WHERE JobID = ?
            GROUP BY StaffName
        """, (job_id,))
        staff_rows = cursor.fetchall()

        users = []
        for staff in staff_rows:
            users.append({
                "name": staff[0],
                "hours": float(staff[1] or 0)
            })

        conn.close()

        return {
            'jobId': job_id,
            'estimatedTime': float(estimated_time or 0),
            'users': users,
            'remainingTime': round(remaining_time, 2),
            'totalHoursWorked': round(total_worked, 2),
            'totalLaborCost': round(total_labor_cost, 2),
            'status': status
        }

    except Exception as e:
        print("Error in get_job_work_details:", str(e))
        return {
            'jobId': job_id,
            'estimatedTime': 0,
            'users': [],
            'remainingTime': 0,
            'totalHoursWorked': 0,
            'totalLaborCost': 0,
            'status': 'Error'
        }


def get_pn_data_details(pn_id):
    """Fetch job details directly from the PN_DATA table."""
    try:
        conn = get_db_connection()

        cursor = conn.cursor()

        # Fetch job details from PN_DATA table for the given PN
        cursor.execute('''
            SELECT 
                [INPUT DATE], PN, [NO/CELL], [DRAW NO], [REQU-DATE], CUST, [STOCK CODE],
                QTY, [CELL CODE], B$, [ORDER NO], MODEL, VOL, AH, WH, CHEM, STRUCTURE, STAFF,
                WORKHR, [HR/PP], [END DATE], [TEST TIME], AV, S$, [C-DRAW], [C-CELLS], [C-AV],
                [C-B$], [C-S$], [C-STCODE], [ORIGINAL S$], DISCOUNT, SALESMAN, [Customer Code], [Order Date],
                [Production Ready Date]
            FROM PN_DATA
            WHERE PN = ?
        ''', (pn_id,))

        row = cursor.fetchone()
        conn.close()

        if row:
            # Map each field to the correct key, guarding against missing fields
            keys = [
                'inputDate', 'pn', 'noCell', 'drawNo', 'requDate', 'cust', 'stockCode',
                'qty', 'cellCode', 'b$', 'orderNo', 'model', 'vol', 'ah', 'wh', 'chem', 'structure', 'staff',
                'workhr', 'hrPerP', 'endDate', 'testTime', 'av', 's$', 'cDraw', 'cCells', 'cAv',
                'cB$', 'cS$', 'cStcode', 'originalS$', 'discount', 'salesman', 'customerCode', 'orderDate',
                'productionReadyDate'
            ]
            # Pad row to match length
            row += tuple([""] * (len(keys) - len(row)))
            return dict(zip(keys, row))

        else:
            return {'error': 'No data found for the provided PN', 'pnId': pn_id}

    except Exception as e:
        print(f"Error in get_pn_data_details: {e}")
        return {'error': str(e), 'pnId': pn_id}

    
#from store_merged_data import store_merged_data_to_db

# Call this function during initialization if required
#store_merged_data_to_db()


def update_testrecord_from_clock(job_id):
    print(f"🔄 [DEBUG] update_testrecord_from_clock called for job_id = {job_id}")
    """Update TestRecords for the given PN from current ClockInOut values."""
    try:
        conn = get_db_connection()

        cursor = conn.cursor()

        # Calculate total_time (excluding QC)
        cursor.execute('''
            SELECT COALESCE(SUM(
                (strftime('%s', COALESCE(StopTime, datetime('now','localtime'))) - strftime('%s', StartTime)) / 3600.0
            ), 0.0)
            FROM ClockInOut
            WHERE JobID = ? AND LOWER(StaffName) != 'qc'
        ''', (job_id,))
        total_time = float(cursor.fetchone()[0] or 0)

        #calculate Remaining Time

        # Calculate test_time (QC only)
        cursor.execute('''
            SELECT COALESCE(SUM(
                (strftime('%s', COALESCE(StopTime, datetime('now','localtime'))) - strftime('%s', StartTime)) / 3600.0
            ), 0.0)
            FROM ClockInOut
            WHERE JobID = ? AND LOWER(StaffName) = 'qc'
        ''', (job_id,))
        test_time = float(cursor.fetchone()[0] or 0)

        # Get rate from config or use default
        hourly_rate = 37.95
        total_labor_cost = round(total_time * hourly_rate, 2)

        # Get qty, av, unit_price, bill_price
        cursor.execute('SELECT qty, av, unit_price, bill_price FROM TestRecords WHERE PN = ?', (job_id,))
        row = cursor.fetchone()
        if row:
            qty, av, unit_price, bill_price = row
            qty = float(qty or 0)
            av = float(av or 0)
            estimated_time = round(qty * av, 2)
            profit = round((qty * float(unit_price or 0)) - total_labor_cost - (qty * float(bill_price or 0)), 2)
            remaining_time = round(estimated_time - total_time, 2)
        else:
            estimated_time = 0
            profit = 0

        # Update TestRecords
        cursor.execute('''
            UPDATE TestRecords SET
                total_time = ?,
                remaining_time = ?,
                test_time = ?,
                total_labor_cost = ?,
                profit = ?,
                estimated_time = ?
            WHERE PN = ?
        ''', (total_time,remaining_time, test_time, total_labor_cost, profit, estimated_time, job_id))

        # Update staff_details
        cursor.execute('''
            SELECT StaffName, ROUND(SUM(
                (strftime('%s', COALESCE(StopTime, datetime('now','localtime'))) - strftime('%s', StartTime)) / 3600.0), 2)
            FROM ClockInOut
            WHERE JobID = ?
            GROUP BY StaffName
        ''', (job_id,))
        staff_rows = cursor.fetchall()
        staff_details = {row[0]: {'worked_hours': row[1]} for row in staff_rows}
        cursor.execute('UPDATE TestRecords SET staff_details = ? WHERE PN = ?', (json.dumps(staff_details), job_id))

        # Update JOBSFINISHED if present
        cursor.execute('SELECT 1 FROM JOBSFINISHED WHERE PN = ?', (job_id,))
        if cursor.fetchone():
            cursor.execute('''
                UPDATE JOBSFINISHED SET
                    TotalHoursWorked = ?,
                    "TEST TIME" = ?,
                    TotalLaborCost = ?,
                    EstimatedTime = ?
                WHERE PN = ?
            ''', (total_time, test_time, total_labor_cost, estimated_time,job_id))

        conn.commit()
    except Exception as e:
        print(f"Error updating TestRecords/JobsFinished for PN={job_id}: {e}")
    finally:
        if 'conn' in locals() and conn:
            conn.close()






# Define the HTTP request handler
class ClockInOutHandler(BaseHTTPRequestHandler):
    def set_cors_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type,X-User-Role')
        self.send_header('Cache-Control', 'no-store')
    
    
    def do_OPTIONS(self):
        # Handle preflight requests for CORS
        self.send_response(200)
        self.set_cors_headers()
        self.send_header('X-Content-Type-Options', 'nosniff')  
        self.end_headers()
    # Add the /view-running-jobs endpoint to the do_GET method
    def do_GET(self):
        #self.send_response(200)
        #self.set_cors_headers()

        if self.path in ["/", "/Sydney_layout", "/Sydney_layout.html"]:
            template = env.get_template("Sydney_layout.html")
            html = template.render()
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(html.encode("utf-8"))
            return

        elif self.path == "/dashboard":
            template = env.get_template("dashboard.html")
            html = template.render()
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(html.encode("utf-8"))
            return

        elif self.path == "/timesheet":
            template = env.get_template("Timesheet.html")
            html = template.render()
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(html.encode("utf-8"))
            return

        elif self.path == "/finished_jobs":
            template = env.get_template("finished-jobs.html")
            html = template.render()
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(html.encode("utf-8"))
            return
        
        elif self.path == "/PN_DATA":
            template = env.get_template("PN_DATA.html")
            html = template.render()
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(html.encode("utf-8"))
            return
        
        elif self.path == "/Login":
            template = env.get_template("Login.html")
            html = template.render()
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(html.encode("utf-8"))
            return
        
        elif self.path.startswith("/formtopdf"):
            template = env.get_template("formtopdf.html")
            html = template.render()
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(html.encode("utf-8"))
            return

        
        elif self.path == "/view_allocated_jobs":
            template = env.get_template("View_allocated_jobs.html")
            html = template.render()
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(html.encode("utf-8"))
            return
        
        elif self.path == "/add_job":
            template = env.get_template("add_job.html")
            html = template.render()
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(html.encode("utf-8"))
            return

        elif self.path == "/Allocate_job":
            template = env.get_template("Allocate_jobs.html")
            html = template.render()
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(html.encode("utf-8"))
            return

        elif self.path == "/Products_test_records":
            template = env.get_template("Products_test_records.html")
            html = template.render()
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(html.encode("utf-8"))
            return    



        elif self.path == '/get-staff':
            try:
                import os
                conn = get_db_connection()

                cursor = conn.cursor()

                cursor.execute("SELECT StaffID, StaffName FROM Staff ORDER BY StaffName;")
                rows = cursor.fetchall()
                print(f">>> DB in use: {os.path.abspath(DB_NAME)} | Staff count: {len(rows)}")

                staff = [{"id": row[0], "name": row[1]} for row in rows]
                conn.close()

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'staff': staff}).encode('utf-8'))
            except Exception as e:
                print(">>> ERROR in /get-staff:", e)
                self.send_response(500)
                self.set_cors_headers()
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))
            return




        elif self.path == '/get-jobs':
            try:
                conn = get_db_connection()

                cursor = conn.cursor()

                cursor.execute("""
                    SELECT 
                        p.PN, p.CUST, p."REQU-DATE", 
                        CASE 
                            WHEN EXISTS (SELECT 1 FROM AllocatedJobs a WHERE a.JobID = p.PN) 
                            THEN 1 ELSE 0 
                        END AS isAssigned,
                        COALESCE(p.backorder, 0) AS backorder
                    FROM PN_DATA p
                """)
                jobs = cursor.fetchall()
                conn.close()

                job_records = [{
                    'jobId': row[0],
                    'customer': row[1],
                    'requiredDate': row[2],
                    'isAssigned': bool(row[3]),
                    'backorder': bool(row[4])
                } for row in jobs]
                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                response = {'jobs': job_records}
                self.wfile.write(json.dumps(response).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))



        elif self.path.startswith('/get-allocated-jobs'):
            try:
                # Parse the query parameters
                parsed_url = urlparse(self.path)
                query_params = parse_qs(parsed_url.query)
                staff_name = query_params.get('staffName', [None])[0]

                if not staff_name:
                    self.send_response(400)
                    self.set_cors_headers()
                    self.end_headers()
                    self.wfile.write(json.dumps({'error': 'Staff name is required'}).encode('utf-8'))
                    return

                # Connect to the database
                conn = get_db_connection()

                cursor = conn.cursor()



                # Query to fetch allocated jobs for the given staff name
                cursor.execute('''
                    SELECT JobID,CustomerName, AllocationDate
                    FROM AllocatedJobs
                    WHERE LOWER(StaffName) = LOWER(?)
                ''', (staff_name,))
                rows = cursor.fetchall()
                conn.close()

                # Format the response
                allocated_jobs = [{'jobId': row[0],'customerName': row[1], 'allocationDate': row[2]} for row in rows]

                # Send the response
                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'jobs': allocated_jobs}).encode('utf-8'))

            except Exception as e:
                print(f"Error fetching allocated jobs: {e}")
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))

        elif self.path.startswith("/get-av"):
            query_string = self.path.split('?')[-1]
            query_params = urllib.parse.parse_qs(query_string)
            stock_code = query_params.get('stockCode', [None])[0]

            if stock_code:
                data = get_av_by_stock_code(stock_code)
                if data:
                    response = {'avValue': data['avValue'], 'drawNo': data['drawNo']}
                else:
                    response = {'error': 'No AV found for the given STOCK CODE'}
            else:
                response = {'error': 'STOCK CODE parameter is required'}

            self.send_response(200)
            self.set_cors_headers()
            self.send_header('Content-Type', 'application/json')
            #self.send_header('Cache-Control', 'no-store') 
            self.end_headers()
            self.wfile.write(json.dumps(response).encode('utf-8'))

            
        elif self.path == '/get-totallaborcost':
            try:
                conn = get_db_connection()

                cursor = conn.cursor()

                # Query to fetch Job IDs and their total labor costs
                cursor.execute('''
                    SELECT JobID, TotalLaborCost FROM JobTable
                ''')
                jobs = cursor.fetchall()
                conn.close()

                # Prepare the response in JSON format
                job_records = [{'jobId': job[0], 'totalLaborCost': job[1] or 0.0} for job in jobs]

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                response = {'jobs': job_records}
                self.wfile.write(json.dumps(response).encode('utf-8'))

            except Exception as e:
                print(f"Error in /get-totallaborcost: {str(e)}")
                self.send_response(500)
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))

            
        




        
        elif self.path.startswith("/view-times"):
            try:
                # Parse query parameters
                parsed_url = urlparse(self.path)
                query_params = parse_qs(parsed_url.query)

                # Pagination
                page = int(query_params.get('page', ['1'])[0])
                page_size = int(query_params.get('page_size', ['10'])[0])
                offset = (page - 1) * page_size

                # Filters
                filter_column = query_params.get('filter_column', ['all'])[0]
                filter_value = query_params.get('filter_value', [''])[0].lower()

                conn = get_db_connection()

                cursor = conn.cursor()

                # --- BASE QUERY (no WHERE here). NOTE: Task added right after JobID ---
                base_query = '''
                SELECT 
                    c.RecordID                          AS RecordID,        -- 0
                    c.StaffName                         AS StaffName,       -- 1
                    c.JobID                             AS JobID,           -- 2
                    c.Task                              AS Task,            -- 3  ✅ NEW
                    c.StartTime                         AS StartTime,       -- 4
                    c.StopTime                          AS StopTime,        -- 5
                    c.LaborCost                         AS LaborCost,       -- 6
                    j.CUST                              AS CustomerName,    -- 7
                    j."DRAW NO"                         AS DrawingNumber,   -- 8
                    j."NO/CELL"                         AS CellNo,          -- 9
                    j.QTY                               AS Quantity,        -- 10
                    j."REQU-DATE"                       AS RequestDate,     -- 11
                    COALESCE(j.AV * j.QTY, 0.0)         AS EstimatedTime,   -- 12
                    ROUND(
                        COALESCE(
                            (strftime('%s', c.StopTime) - strftime('%s', c.StartTime)) / 3600.0, 
                            0.0
                        ), 2
                    )                                   AS TotalHoursWorked, -- 13
                    CASE 
                    WHEN jt.Status = 'Finished'  THEN 'Finished'
                    WHEN f.Status  = 'Completed' THEN 'Completed'
                    ELSE 'Active'
                    END                                 AS Status           -- 14
                FROM ClockInOut c
                LEFT JOIN PN_DATA      j  ON c.JobID = j.PN
                LEFT JOIN JOBSFINISHED f  ON c.JobID = f.PN
                LEFT JOIN JobTable     jt ON c.JobID = jt.JobID
                '''

                # --- Build OUTER where on aliased columns (wrap the base query as q) ---
                column_map = {
                    '0':  'q.StaffName',
                    '1':  'q.JobID',
                    # if you add a Task filter option in UI, map its index here, e.g. '12': 'q.Task'
                    '2':  'q."DrawingNumber"',
                    '3':  'q."CellNo"',
                    '4':  'q.Quantity',
                    '5':  'q.CustomerName',
                    '6':  'q.StartTime',
                    '7':  'q.StopTime',
                    '8':  'q.TotalHoursWorked',
                    '9':  'q.EstimatedTime',
                    '10': '(q.EstimatedTime - q.TotalHoursWorked)',
                    '11': 'q.LaborCost'
                }

                where_clauses, params = [], []
                if filter_value and filter_column != 'all' and filter_column in column_map:
                    where_clauses.append(f"LOWER({column_map[filter_column]}) LIKE ?")
                    params.append(f"%{filter_value}%")
                where_stmt = (' WHERE ' + ' AND '.join(where_clauses)) if where_clauses else ''

                wrapped_query = f"SELECT * FROM ({base_query}) AS q"

                # COUNT
                count_query = f"SELECT COUNT(*) FROM ({base_query}) AS q{where_stmt}"
                cursor.execute(count_query, params)
                total_records = cursor.fetchone()[0]

                # DATA
                data_query = f"""
                    {wrapped_query}
                    {where_stmt}
                    ORDER BY q.StartTime DESC
                    LIMIT ? OFFSET ?
                """
                cursor.execute(data_query, params + [page_size, offset])
                rows = cursor.fetchall()
                conn.close()

                # --- Build JSON (indexes updated for Task) ---
                records = []
                for row in rows:
                    record = {
                        'recordId':         row[0],
                        'staffName':        row[1],
                        'jobId':            row[2],
                        'task':             row[3] or 'N/A',     # ✅ Task
                        'startTime':        row[4] or 'NA',
                        'stopTime':         row[5] if row[5] else "In Progress",
                        'laborCost':        row[6] if row[6] is not None else 0.0,
                        'customerName':     row[7] or " ",
                        'drawingNumber':    row[8] or " ",
                        'cellNo':           row[9] or " ",
                        'quantity':         row[10] or " ",
                        'requDate':         row[11],
                        'estimatedTime':    float(row[12] or 0.0),
                        'totalHoursWorked': float(row[13] or 0.0),
                        'remainingTime':    round(max(float(row[12] or 0.0) - float(row[13] or 0.0), 0.0), 2),
                        'status':           row[14]
                    }
                    records.append(record)

                total_pages = (total_records + page_size - 1) // page_size
                response = {
                    'records': records,
                    'totalRecords': total_records,
                    'totalPages': total_pages,
                    'currentPage': page
                }

                # Compressed response
                json_data = json.dumps(response).encode('utf-8')
                buffer = BytesIO()
                with gzip.GzipFile(fileobj=buffer, mode='wb') as gzip_file:
                    gzip_file.write(json_data)

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.send_header('Content-Encoding', 'gzip')
                self.send_header('Content-Length', str(len(buffer.getvalue())))
                self.end_headers()
                self.wfile.write(buffer.getvalue())

            except Exception as e:
                print(f"Error in /view-times: {str(e)}")
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))



        elif self.path == '/get-pn-data':
            conn = None
            try:
                conn = get_db_connection()

                cursor = conn.cursor()
                
                # List of column names you want to hide
                ignore_columns = ["C-DRAW", "C-CELLS", "C-AV", "C-B$", "C-S$", "C-STCODE","DISCOUNT","WORKHR","HR/PP","ORIGINAL S$","TEST TIME"]

                # Fetch column info
                cursor.execute("PRAGMA table_info(PN_DATA)")
                columns_info = cursor.fetchall()

                if not columns_info:
                    self.send_response(404)
                    self.set_cors_headers()
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'error': 'PN_DATA table not found',
                        'tables': [row[0] for row in cursor.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
                    }).encode())
                    return

                # Filter out unwanted columns
                columns = [col[1] for col in columns_info if col[1] not in ignore_columns]

                # Build the SELECT query safely
                columns_str = ', '.join(f'"{col}"' for col in columns)
                query = f'SELECT {columns_str} FROM PN_DATA ORDER BY "INPUT DATE" DESC'

                # Execute query
                try:
                    cursor.execute(query)
                    pn_data = cursor.fetchall()
                except sqlite3.Error as e:
                    self.send_response(500)
                    self.set_cors_headers()
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'error': f'Database query failed: {str(e)}',
                        'query': query
                    }).encode())
                    return

                # Convert to list of dictionaries
                data = []
                for row in pn_data:
                    data.append({columns[i]: str(value) if value is not None else '' 
                                for i, value in enumerate(row)})

                # Send success response
                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(data).encode('utf-8'))

            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({
                    'error': f'Failed to fetch PN_DATA: {str(e)}',
                    'type': type(e).__name__
                }).encode('utf-8'))
            finally:
                if conn:
                    conn.close()



        elif self.path.startswith('/view-running-jobs'):
            
            try:
                parsed = urlparse(self.path)
                q = parse_qs(parsed.query)
                task = q.get('task', [None])[0]  # e.g. /view-running-jobs?task=Welding

                conn = get_db_connection()

                cursor = conn.cursor()

                sql = """
                    SELECT c.StaffName, c.JobID, j.CUST, j."DRAW NO", c.StartTime, c.Task
                    FROM ClockInOut c
                    LEFT JOIN PN_DATA j ON c.JobID = j.PN
                    WHERE c.StopTime IS NULL
                """
                params = []
                if task:
                    sql += " AND c.Task = ?"
                    params.append(task)

                sql += " ORDER BY c.StartTime DESC"
                cursor.execute(sql, params)

                rows = cursor.fetchall()
                conn.close()
                print(f"Running jobs fetched (task={task}): {rows[:3]} ... total={len(rows)}")

                running_jobs = [{
                    'staffName':   r[0],
                    'jobId':       r[1],
                    'customerName':r[2] or 'N/A',
                    'drawNumber':  r[3] or 'N/A',
                    'startTime':   r[4],
                    'task':        r[5] or 'N/A',   # 👈 include task in response
                } for r in rows]

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'runningJobs': running_jobs}).encode('utf-8'))

            except Exception as e:
                print(f"Error fetching running jobs: {str(e)}")
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))


        elif self.path == '/export-to-excel':
            # Export the database table to Excel and send it to the client
            try:
                output_file = 'clock_in_data.xlsx'
                query = "SELECT * FROM ClockInOut"
                connection = sqlite3.connect('prod_management.db')
                df = pd.read_sql_query(query, connection)
                
                # Save the Excel file to a BytesIO stream
                output_stream = io.BytesIO()
                df.to_excel(output_stream, index=False, engine='openpyxl')
                output_stream.seek(0)  # Rewind the stream

                # Send the Excel file as a response
                self.send_response(200)
                self.send_cors_headers()
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                #self.send_header('Cache-Control', 'no-store')
                self.send_header('Content-Disposition', 'attachment; filename="clock_in_data.xlsx"')
                self.end_headers()
                self.wfile.write(output_stream.read())  # Write the file content to the response

            except Exception as e:
                # Handle errors
                self.send_response(500)
                self.set_cors_headers()
                #self.send_header('Cache-Control', 'no-store')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))

        

        elif self.path.startswith('/view-finished-jobs'):
            try:
                parsed_url = urlparse(self.path)
                query_params = parse_qs(parsed_url.query)

                page = int(query_params.get('page', ['1'])[0])
                page_size = int(query_params.get('page_size', ['10'])[0])
                search_term = query_params.get('custName', [''])[0]
                offset = (page - 1) * page_size

                conn = get_db_connection()

                cursor = conn.cursor()

                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='JOBSFINISHED'")
                table_exists = cursor.fetchone()

                if not table_exists:
                    response = {'jobs': [], 'total_pages': 0, 'current_page': 1}
                else:
                    # List of column names to ignore
                    ignore_columns = ["C-DRAW", "C-CELLS", "C-AV", "C-B$", "C-S$", "C-STCODE","WORKHR","HR/PP","ORIGINAL S$","DISCOUNT"]

                    # Get all columns
                    cursor.execute("PRAGMA table_info(JOBSFINISHED)")
                    columns_info = cursor.fetchall()
                    columns = [col[1] for col in columns_info if col[1] not in ignore_columns]

                    if search_term:
                        like_pattern = f"%{search_term}%"
                        cursor.execute("SELECT COUNT(*) FROM JOBSFINISHED WHERE CUST LIKE ?", (like_pattern,))
                        total_records = cursor.fetchone()[0]
                        total_pages = (total_records + page_size - 1) // page_size

                        columns_str = ', '.join(f'"{col}"' for col in columns)
                        query = f"SELECT {columns_str} FROM JOBSFINISHED WHERE CUST LIKE ? LIMIT ? OFFSET ?"
                        cursor.execute(query, (like_pattern, page_size, offset))
                    else:
                        cursor.execute("SELECT COUNT(*) FROM JOBSFINISHED")
                        total_records = cursor.fetchone()[0]
                        total_pages = (total_records + page_size - 1) // page_size

                        columns_str = ', '.join(f'"{col}"' for col in columns)
                        query = f"SELECT {columns_str} FROM JOBSFINISHED ORDER BY COALESCE(\"END DATE\", '') DESC LIMIT ? OFFSET ?"

                        cursor.execute(query, (page_size, offset))

                    rows = cursor.fetchall()

                    finished_jobs = [
                        {columns[i]: value for i, value in enumerate(row)}
                        for row in rows
                    ]

                    response = {
                        'jobs': finished_jobs,
                        'total_pages': total_pages,
                        'current_page': page,
                        'total_records': total_records
                    }

                conn.close()

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode('utf-8'))

            except Exception as e:
                print(f"Error fetching finished jobs: {str(e)}")
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))



        elif self.path == "/api/test-records":
            self.handle_get_test_records()
            return


        elif self.path.startswith("/get-job-work-details"):
            query_string = self.path.split('?')[-1]
            query_params = urllib.parse.parse_qs(query_string)
            job_id = query_params.get('jobId', [None])[0]

            if not job_id:
                self.send_response(400)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': 'Job ID is required'}).encode('utf-8'))
                return

            # Get work details for job
            job_details = get_job_work_details(job_id)
            self.send_response(200)
            self.set_cors_headers()
            self.end_headers()
            self.wfile.write(json.dumps(job_details).encode('utf-8'))

            
        
        
        elif self.path.startswith('/get-csv-data'):
            try:
                # Parse query parameters
                query = urlparse(self.path).query
                params = parse_qs(query)
                Drawing_Number = params.get('Drawing_Number', [None])[0]
                print(f"🔍 Received Drawing_Number: {Drawing_Number} (Type: {type(Drawing_Number)})")


                conn = get_db_connection()

                cursor = conn.cursor()

                # Build query based on file_name
                if Drawing_Number and Drawing_Number.strip():
                    Drawing_Number = Drawing_Number.strip()  # Trim spaces
                    sql_query = 'SELECT * FROM csv_data WHERE Drawing_Number = ? ORDER BY DATE DESC'
                    print(f"📝 Executing SQL: {sql_query} with value '{Drawing_Number}'")
                    cursor.execute(sql_query, (Drawing_Number,))
                else:
                    print("⚠️ Drawing_Number is missing, returning all records.")
                    cursor.execute('SELECT * FROM csv_data')

                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()
                conn.close()


               

                # Convert rows to list of dictionaries
                csv_data = [dict(zip(columns, row)) for row in rows]
                hidden_column = 'NEW'  # Column to hide
                for row in csv_data:
                    if hidden_column in row:
                        del row[hidden_column] 

                
                for i, row in enumerate(csv_data):
                    if i == 0:
                        # First row: Set TOTAL_AV = AVERAGE_TIME
                        row['TOTAL_AV'] = row['AVERAGE_TIME']
                    else:
                        # Subsequent rows: Set TOTAL_AV to empty string
                        row['TOTAL_AV'] = ' '
                    

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'csvData': csv_data}).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))



        elif self.path == "/get-config":
             # Serve the BASE_URL to the frontend
            self.send_response(200)
            self.set_cors_headers()
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            response = {'base_url': BASE_URL}
            self.wfile.write(json.dumps(response).encode('utf-8'))




        elif self.path.startswith('/get-job-data-by-drawing'):
            try:
                parsed_url = urlparse(self.path)
                query_params = parse_qs(parsed_url.query)
                drawing_number = query_params.get('Drawing_Number', [None])[0]

                conn = get_db_connection()

                cursor = conn.cursor()
                
                # Get latest csv_data
                cursor.execute('''
                    SELECT AVERAGE_TIME, B_PRICE, S_PRICE 
                    FROM csv_data 
                    WHERE Drawing_Number = ?
                    ORDER BY DATE DESC 
                    LIMIT 1
                ''', (drawing_number,))
                
                csv_data = cursor.fetchone()
                result = {'success': False}
                
                if csv_data:
                    result['success'] = True
                    result['csvData'] = {
                        'AVERAGE_TIME': csv_data[0],
                        'B_PRICE': csv_data[1],
                        'S_PRICE': csv_data[2]
                    }
                    
                    # Step 2: Get corresponding stock info from MergedData
                    cursor.execute('''
                        SELECT STOCKCODE, CELLS_PARTS, MODEL 
                        FROM MergedData 
                        WHERE DrawNo = ?
                        LIMIT 1
                    ''', (drawing_number,))
                    
                    merged_row = cursor.fetchone()
                    if merged_row:
                        result['stockCode'] = merged_row[0] or "N/A"
                        result['cellsParts'] = merged_row[1] or "N/A"
                        result['model'] = merged_row[2] or "N/A"

                    
                
                conn.close()

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(result).encode())

            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': False,
                    'error': str(e)
                }).encode())




        elif self.path.startswith('/get-pdf'):
            try:
                # Parse the query parameters
                parsed_url = urlparse(self.path)
                query_params = parse_qs(parsed_url.query)
                pn = query_params.get('pn', [None])[0]

                if not pn:
                    self.send_response(400)
                    self.set_cors_headers()
                    self.end_headers()
                    self.wfile.write(b'PN (part number) is required')
                    return

                # Construct the file path
                file_path = os.path.join(PRODUCT_PHOTOS, f"{pn}.pdf")

                # Check if the file exists
                if not os.path.isfile(file_path):
                    self.send_response(404)
                    self.set_cors_headers()
                    self.end_headers()
                    self.wfile.write(b'PDF not found')
                    return
                
                # Get file size for Content-Length header
                file_size = os.path.getsize(file_path)

                # Serve the PDF file
                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/pdf')
                self.send_header('Content-Disposition', f'inline; filename="{pn}.pdf"')
                self.send_header('Content-Length', str(file_size))
                self.end_headers()
                
                
                with open(file_path, 'rb') as f:
                    while True:
                        chunk = f.read(8192)  # 8KB chunks
                        if not chunk:
                            break
                        self.wfile.write(chunk)

            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(f"Error: {str(e)}".encode())

        elif self.path.startswith('/has-clockinout-records'):
            try:
                parsed_url = urlparse(self.path)
                query_params = parse_qs(parsed_url.query)
                job_id = query_params.get('jobId', [None])[0]

                if not job_id:
                    self.send_response(400)
                    self.set_cors_headers()
                    self.end_headers()
                    self.wfile.write(json.dumps({'error': 'Job ID is required'}).encode())
                    return

                conn = get_db_connection()

                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM ClockInOut WHERE JobID = ?", (job_id,))
                count = cursor.fetchone()[0]
                conn.close()

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'hasRecords': count > 0}).encode())
            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode())



        elif self.path.startswith('/has-clockinout-records'):
            try:
                parsed_url = urlparse(self.path)
                query_params = parse_qs(parsed_url.query)
                job_id = query_params.get('jobId', [None])[0]

                if not job_id:
                    self.send_response(400)
                    self.set_cors_headers()
                    self.end_headers()
                    self.wfile.write(json.dumps({'error': 'Job ID is required'}).encode())
                    return

                conn = get_db_connection()

                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM ClockInOut WHERE JobID = ?", (job_id,))
                count = cursor.fetchone()[0]
                conn.close()

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'hasRecords': count > 0}).encode())
            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode())


        # In do_GET method
        elif self.path.startswith('/refresh-pdf-status'):
            try:
                pn = self.path.split('?pn=')[1]
                file_path = os.path.join(PRODUCT_PHOTOS, f"{pn}.pdf")
                
                conn = get_db_connection()

                cursor = conn.cursor()
                exists = os.path.exists(file_path)
                
                cursor.execute('''
                    INSERT OR REPLACE INTO TEST_RECORDS 
                    (pn, file_path) 
                    VALUES (?, ?)
                ''', (pn, file_path if exists else None))
                
                conn.commit()
                conn.close()
                
                self.send_response(200)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({exists: exists}).encode())
                
            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode())

        elif self.path.startswith("/get-full-job-data"):
            parsed_url = urlparse(self.path)
            query_params = parse_qs(parsed_url.query)
            job_id = query_params.get('jobId', [None])[0]

            if not job_id:
                self.send_response(400)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': 'Job ID is required'}).encode())
                return

            try:
                conn = get_db_connection()

                cursor = conn.cursor()

                # --- Determine Job Status ---
                status = "Not Started"
                cursor.execute("SELECT 1 FROM TestRecords WHERE PN = ? LIMIT 1", (job_id,))
                is_completed = cursor.fetchone() is not None
                if is_completed:
                    status = "Completed"
                else:
                    cursor.execute("SELECT 1 FROM ClockInOut WHERE JobID = ? LIMIT 1", (job_id,))
                    if cursor.fetchone():
                        status = "In Progress"

                # --- Fetch job data ---
                if status == "Completed":
                    cursor.execute("""
                        SELECT PN, draw_no, customer, qty, stock_code, salesman, order_no, 
                            unit_price AS s$, bill_price AS b$, date AS orderDate
                        FROM TestRecords
                        WHERE PN = ?
                        LIMIT 1
                    """, (job_id,))
                    row = cursor.fetchone()
                    pn_data = {
                        "pn": row[0],
                        "drawNo": row[1],
                        "cust": row[2],
                        "qty": row[3],
                        "stockCode": row[4],
                        "salesman": row[5],
                        "orderNo": row[6],
                        "s$": row[7],
                        "b$": row[8],
                        "orderDate": row[9],
                    } if row else {}
                else:
                    pn_data = get_pn_data_details(job_id)

                drawing_no = pn_data.get("drawNo") or pn_data.get("DRAWNO") or ""
                qty = float(pn_data.get("qty") or 0)
                s_price = float(pn_data.get("s$") or 0)
                b_price = float(pn_data.get("b$") or 0)

                # --- Work details ---
                if is_completed:
                    cursor.execute("""
                        SELECT estimated_time, total_time, remaining_time, total_labor_cost, profit, staff_details
                        FROM TestRecords
                        WHERE PN = ?
                        LIMIT 1
                    """, (job_id,))
                    row = cursor.fetchone()
                else:
                    row = None

                if row:
                    estimated_time = float(row[0] or 0)
                    total_time = float(row[1] or 0)
                    remaining_time = float(row[2] or 0)
                    total_labor = float(row[3] or 0) if status == "Completed" else 0
                    profit = float(row[4] or 0) if status == "Completed" else 0
                    staff_json = row[5] or "{}"
                else:
                    # For in-progress / not started
                    try:
                        av = float(pn_data.get("av") or pn_data.get("AV") or 0)
                        estimated_time = round(av * qty, 2)
                    except Exception:
                        estimated_time = 0

                    cursor.execute('''
                        SELECT COALESCE(SUM(
                            CASE 
                                WHEN (julianday(COALESCE(StopTime, datetime('now', 'localtime'))) - julianday(StartTime)) * 24.0 < 0 THEN 0
                                ELSE (julianday(COALESCE(StopTime, datetime('now', 'localtime'))) - julianday(StartTime)) * 24.0
                            END
                        ), 0.0)
                        FROM ClockInOut WHERE JobID = ?
                    ''', (job_id,))
                    total_time = round(float(cursor.fetchone()[0] or 0), 2)
                    remaining_time = round(estimated_time - total_time, 2)
                    total_labor = 0
                    profit = 0
                    staff_json = "{}"

                # --- Parse Staff Details ---
                users = []
                try:
                    staff_dict = json.loads(staff_json or "{}")
                    for k, v in staff_dict.items():
                        try:
                            # Ignore save_time entries
                            if k.endswith("_(save_time)"):
                                continue

                            clean_name = (
                                k.replace("_(hours_worked)", "")
                                .replace("_(save_time)", "")
                                .strip()
                            )
                            if clean_name.lower().startswith("[object_object]") or clean_name == "":
                                continue

                            hours = 0.0
                            if isinstance(v, (int, float)):
                                hours = float(v)
                            elif isinstance(v, str) and v.strip() != "":
                                try:
                                    hours = float(v)
                                except ValueError:
                                    hours = 0.0
                            elif isinstance(v, dict):
                                hours = float(v.get("worked_hours", 0) or 0)

                            if hours > 0:
                                users.append({"name": clean_name, "hours": hours})
                        except Exception as inner:
                            print(f"⚠️ Failed to process staff {k}: {v} → {inner}")
                except Exception as e:
                    print("❌ Failed to parse staff_json:", staff_json, "→", e)

                # --- Fetch CSV Data (safe) ---
                csv_data = []
                if drawing_no:
                    try:
                        cursor.execute('SELECT * FROM csv_data WHERE "Drawing_Number" = ? AND PN = ?', (drawing_no, job_id))
                        columns = [desc[0] for desc in cursor.description]
                        rows = cursor.fetchall()
                        csv_data = [dict(zip(columns, row)) for row in rows]
                    except Exception as e:
                        print(f"⚠️ CSV fetch failed for job {job_id}: {e}")

                # --- Add ClockInOut logs (include Task) ---
                cursor.execute('''
                    SELECT StaffName, StartTime, StopTime, Task,
                        ROUND((julianday(COALESCE(StopTime, datetime('now', 'localtime'))) - julianday(StartTime)) * 24.0, 2) as hours
                    FROM ClockInOut
                    WHERE JobID = ?
                    ORDER BY StartTime
                ''', (job_id,))
                log_rows = cursor.fetchall()
                logs = [{
                    "staffName": row[0],
                    "startTime": row[1],
                    "stopTime": row[2] or "In Progress",
                    "task": row[3] or "N/A",
                    "hours": row[4]
                } for row in log_rows]

                conn.close()

                # --- Build Response ---
                response = {
                    "pnData": {
                        "pn": job_id,
                        "cust": pn_data.get("cust", ""),
                        "drawNo": drawing_no,
                        "stockCode": pn_data.get("stockCode", ""),
                        "qty": qty,
                        "salesman": pn_data.get("salesman", ""),
                        "orderDate": pn_data.get("orderDate", ""),
                        "s$": s_price,
                        "b$": b_price
                    },
                    "workData": {
                        "estimatedTime": estimated_time,
                        "totalHoursWorked": total_time,
                        "remainingTime": remaining_time,
                        "totalLaborCost": total_labor,
                        "status": status,
                        "users": users,
                        "profit": profit
                    },
                    "csvData": csv_data,
                    "logData": logs
                }

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())

            except Exception as e:
                print(f"❌ Error in /get-full-job-data for {job_id}: {e}")
                self.send_response(500)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"error": str(e)}).encode())
            finally:
                if 'conn' in locals():
                    conn.close()


        elif self.path.startswith("/static/"):
            # Remove query parameters (?v=2.1)
            clean_path = self.path.split("?")[0]
            file_path = clean_path.lstrip("/")  # remove leading slash
            try:
                with open(file_path, "rb") as f:
                    if file_path.endswith(".css"):
                        self.send_response(200)
                        self.send_header("Content-type", "text/css")
                    elif file_path.endswith(".js"):
                        self.send_response(200)
                        self.send_header("Content-type", "application/javascript")
                    elif file_path.endswith(".png"):
                        self.send_response(200)
                        self.send_header("Content-type", "image/png")
                    elif file_path.endswith(".jpg") or file_path.endswith(".jpeg"):
                        self.send_response(200)
                        self.send_header("Content-type", "image/jpeg")
                    else:
                        self.send_response(200)
                        self.send_header("Content-type", "application/octet-stream")
                    self.end_headers()
                    self.wfile.write(f.read())
            except FileNotFoundError:
                self.send_error(404, f"File not found: {file_path}")




        


        elif self.path == '/get-all-job-ids':
            try:
                conn = get_db_connection()

                cursor = conn.cursor()

                # Pull PN and customer (aliased from CUST in PN_DATA)
                cursor.execute('''
                    SELECT PN, CUST AS customer FROM PN_DATA
                    UNION
                    SELECT PN, customer FROM TestRecords
                ''')
                rows = cursor.fetchall()
                conn.close()

                suggestions = []
                for row in rows:
                    pn = row[0]
                    cust = row[1] or ""
                    label = f"{pn} - {cust}" if cust else pn
                    suggestions.append({"jobId": pn, "label": label})

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'suggestions': suggestions}).encode())

            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode())
            return


        elif self.path == '/get-job-status-summary':
            try:
                conn = get_db_connection()

                cursor = conn.cursor()

                # ✅ Total = unique PNs from PN_DATA and TestRecords
                cursor.execute("""
                    SELECT COUNT(DISTINCT PN) FROM (
                        SELECT PN FROM PN_DATA
                        UNION
                        SELECT PN FROM TestRecords
                    ) AS all_jobs
                """)

                total = cursor.fetchone()[0]

                # ✅ Completed jobs = distinct PNs from TestRecords
                cursor.execute("SELECT COUNT(DISTINCT PN) FROM TestRecords")
                completed = cursor.fetchone()[0]

                # ✅ In-progress jobs = in ClockInOut but not in JOBSFINISHED
                cursor.execute("""
                    SELECT COUNT(DISTINCT ClockInOut.JobID)
                    FROM ClockInOut
                    LEFT JOIN JOBSFINISHED ON ClockInOut.JobID = JOBSFINISHED.PN
                    WHERE JOBSFINISHED.PN IS NULL
                """)
                in_progress = cursor.fetchone()[0]

                not_started = max(total - in_progress - completed, 0)

                conn.close()

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'total': total,
                    'inProgress': in_progress,
                    'completed': completed,
                    'notStarted': not_started
                }).encode())

            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode())


        elif self.path == '/get-not-started-jobs':
            try:
                conn = get_db_connection()

                cursor = conn.cursor()

                cursor.execute("""
                    SELECT DISTINCT PN_DATA.PN, PN_DATA.CUST
                    FROM PN_DATA
                    LEFT JOIN ClockInOut ON PN_DATA.PN = ClockInOut.JobID
                    LEFT JOIN TestRecords ON PN_DATA.PN = TestRecords.PN
                    WHERE ClockInOut.JobID IS NULL AND TestRecords.PN IS NULL
                """)
                rows = cursor.fetchall()
                conn.close()

                jobs = [{'jobId': row[0], 'customer': row[1]} for row in rows]

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'jobs': jobs}).encode())
            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode())

        elif self.path == "/get-in-progress-jobs":
            try:
                conn = get_db_connection()

                cursor = conn.cursor()

                cursor.execute("""
                    SELECT c.JobID, j.CUST
                    FROM ClockInOut c
                    LEFT JOIN PN_DATA j ON c.JobID = j.PN
                    WHERE c.StopTime IS NOT NULL
                    AND c.JobID NOT IN (SELECT PN FROM JOBSFINISHED)
                    AND c.JobID NOT IN (SELECT PN FROM TestRecords)
                    GROUP BY c.JobID

                """)
                rows = cursor.fetchall()
                conn.close()

                jobs = [{'jobId': row[0], 'customerName': row[1] or "N/A"} for row in rows]

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'jobs': jobs}).encode())

            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode())










        
        else:
            self.send_response(404)
            self.set_cors_headers()
            #self.send_header('Cache-Control', 'no-store')  #Prevents caching
            self.end_headers()


        

    

        
    
    def do_POST(self):
        '''self.send_response(200)
        self.set_cors_headers()
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')  # Allow all origins
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('X-Content-Type-Options', 'nosniff')'''
        
        if self.path == '/clock-in':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)

            data = json.loads(post_data.decode('utf-8'))
            staff_name = data['staffName']
            job_id = data['jobId']
            task = data['task'] 
            start_time_str = data['startTime']
            stop_time_str = data['stopTime']

            conn = get_db_connection()

            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO ClockInOut (StaffName, JobID, Task, StartTime, StopTime)
                VALUES (?, ?, ?, ?, ?)
            ''', (staff_name, job_id, task, start_time_str, stop_time_str))

            conn.commit()
            conn.close()

            self.send_response(200)
            self.set_cors_headers()
            self.send_header('Content-type', 'application/json')
            self.end_headers()

            response = {'message': 'Data saved successfully!'}
            self.wfile.write(json.dumps(response).encode('utf-8'))



        elif self.path == '/start-job':
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                data = json.loads(post_data.decode('utf-8'))

                staff_name = data['staffName']
                job_id     = data['jobId']
                task       = data.get('task')   # 👈 capture task
                start_time = get_current_timestamp()

                print("➡️ /start-job called")
                print(f"   staffName={staff_name}, jobId={job_id}, task={task}, start_time={start_time}")


                conn = get_db_connection()

                cursor = conn.cursor()

                # 1. Check if staff already has an active job
                cursor.execute('''
                    SELECT JobID FROM ClockInOut
                    WHERE StaffName = ? AND StopTime IS NULL
                ''', (staff_name,))
                active = cursor.fetchone()

                if active:
                    # Respond with ACTIVE_JOB_EXISTS
                    self.send_response(409)  # conflict
                    self.set_cors_headers()
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'code': 'ACTIVE_JOB_EXISTS',
                        'error': f"{staff_name} already working on job {active[0]}"
                    }).encode('utf-8'))
                    return

                # 2. Insert new job with Task
                cursor.execute('''
                    INSERT INTO ClockInOut (StaffName, JobID, Task, StartTime, StopTime)
                    VALUES (?, ?, ?, ?, NULL)
                ''', (staff_name, job_id, task, start_time))

                conn.commit()

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'message': 'Job started successfully!'}).encode('utf-8'))
                

            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))
                
            finally:
                if conn:
                    conn.close()



        elif self.path == '/stop-job':
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                data = json.loads(post_data.decode('utf-8'))
                print(f"Received stop-job data: {data}")  # Log incoming data

                staff_name = data['staffName']
                job_id = data['jobId']
                stop_time = get_current_timestamp()  # Get the current timestamp

                with sqlite3.connect(DB_NAME) as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        UPDATE ClockInOut
                        SET StopTime = ?
                        WHERE StaffName = ? AND JobID = ? AND StopTime IS NULL
                    ''', (stop_time, staff_name, job_id))
                    conn.commit()
                    print(f"Rows updated: {cursor.rowcount}")  # Log row count

                    if cursor.rowcount == 0:
                        raise ValueError("No active job found for the given staffName and jobId.")

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'message': 'Job stopped successfully!'}).encode('utf-8'))

            except Exception as e:
                print(f"Error in /stop-job: {e}")  # Log the error
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))


        elif self.path == '/add-staff':
            try:
                # Read and parse incoming JSON data
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                data = json.loads(post_data.decode('utf-8'))

                # Get staffName
                staff_name = data.get('staffName', '').strip()
                if not staff_name:
                    raise ValueError("Staff Name is required.")

                # Insert into Users table
                conn = get_db_connection()
  # use your DB file name
                cursor = conn.cursor()
                cursor.execute("INSERT INTO Staff (staffName) VALUES (?)", (staff_name,))
                conn.commit()
                conn.close()

                # Respond success
                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'message': 'Staff added successfully!'}).encode('utf-8'))

            except sqlite3.IntegrityError:
                # Staff already exists (UNIQUE constraint failed)
                self.send_response(400)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': 'Staff already exists!'}).encode('utf-8'))

            except Exception as e:
                self.send_response(400)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))

    

        elif self.path == '/delete-staff':
            try:
                # Read and parse JSON data
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                data = json.loads(post_data.decode('utf-8'))

                staff_name_to_delete = data.get('staffName', '').strip()
                if not staff_name_to_delete:
                    raise ValueError("Staff Name is required for deletion.")

                # Connect to DB
                conn = get_db_connection()

                cursor = conn.cursor()

                # Delete from Staff table
                cursor.execute("DELETE FROM Staff WHERE staffName = ?", (staff_name_to_delete,))
                conn.commit()
                rows_deleted = cursor.rowcount
                conn.close()

                if rows_deleted == 0:
                    # No staff found with that name
                    self.send_response(404)
                    self.set_cors_headers()
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({'error': f'Staff "{staff_name_to_delete}" not found'}).encode('utf-8'))
                else:
                    self.send_response(200)
                    self.set_cors_headers()
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({'message': f'Staff "{staff_name_to_delete}" deleted successfully'}).encode('utf-8'))

            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))




        elif self.path == '/add_job':
            try:
                # Read and parse the incoming JSON data
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                data = json.loads(post_data.decode('utf-8'))
                

                # Extract job details from request
                job_id = data.get('pn', '').strip()
                input_date = datetime.now().strftime('%Y-%m-%d')  # Current date
                no_cell = data.get('noCell', '')
                draw_no = data.get('drawNo', '')
                req_date = (datetime.now() + timedelta(weeks=3)).strftime('%Y-%m-%d')  # 3 weeks from input date
                cust = data.get('cust', '')
                stock_code = data.get('stockCode', '')
                qty = data.get('qty', 0)
                cell_code = data.get('cellCode', '')
                b_price = data.get('bPrice', 0)
                order_no = data.get('orderNo', '')
                model = data.get('model', '')
                vol = data.get('vol', 0)
                ah = data.get('ah', 0)
                wh = data.get('wh', 0)
                chem = data.get('chem', '')
                structure = data.get('structure', '')
                staff = data.get('staff', '')
                workhr = data.get('workhr', 0)
                hrpp = data.get('HRPP', 0)
                end_date = data.get('endDate', '')
                test_time = data.get('testTime', 0)
                av = data.get('av', 0)
                s_price = data.get('sPrice', 0)
                discount = data.get('discount', 0)
                salesman = data.get('salesman', '')
                customer_code = data.get('customerCode', '')
                order_date = data.get('orderDate', '')
                exclude_save_time = 1 if data.get('excludeSaveTime') else 0


                # Insert into database
                conn = get_db_connection()

                cursor = conn.cursor()

                try:
                    cursor.execute('''
                        INSERT INTO PN_DATA (PN, "INPUT DATE", "NO/CELL", "DRAW NO", "REQU-DATE", "CUST", 
                                            "STOCK CODE", "QTY", "CELL CODE", "B$", "ORDER NO", "MODEL", "VOL", 
                                            "AH", "WH", "CHEM", "STRUCTURE", "STAFF", "WORKHR", "HR/PP", "END DATE", 
                                            "TEST TIME", "AV", "S$", "DISCOUNT", "SALESMAN", "Customer Code", "Order Date","EXCLUDE_SAVE_TIME") 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?)
                    ''', (job_id, input_date, no_cell, draw_no, req_date, cust, stock_code, qty, cell_code, b_price, 
                        order_no, model, vol, ah, wh, chem, structure, staff, workhr, hrpp, end_date, test_time, 
                        av, s_price, discount, salesman, customer_code, order_date,exclude_save_time))
                    conn.commit()

                    # Send success response
                    self.send_response(200)
                    self.set_cors_headers()
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({'message': 'Job added successfully!'}).encode('utf-8'))

                except sqlite3.IntegrityError as e:
                    if 'UNIQUE constraint failed' in str(e):
                        # Handle duplicate PN error
                        self.send_response(400)
                        self.set_cors_headers()
                        self.send_header('Content-Type', 'application/json')
                        self.end_headers()
                        self.wfile.write(json.dumps({'error': 'The Production Number (PN) already exists. Please use a unique PN.'}).encode('utf-8'))
                    else:
                        raise

                finally:
                    conn.close()

            except Exception as e:
                # Log the error for debugging
                print(f"Error: {e}")
                self.send_response(500)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))
      
      
        elif self.path == '/delete-job':
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                data = json.loads(post_data.decode('utf-8'))

                job_id_to_delete = data.get('jobId', '').strip()
                if not job_id_to_delete:
                    raise ValueError("Job ID is required for deletion.")

                # Database connection and deletion
                conn = get_db_connection()
  # Replace with your actual DB connection
                cursor = conn.cursor()

                # First check if job exists
                cursor.execute("SELECT COUNT(*) FROM PN_DATA WHERE PN = ?", (job_id_to_delete,))
                if cursor.fetchone()[0] == 0:
                    raise ValueError("Job ID not found in database")

                # Delete the job
                cursor.execute("DELETE FROM PN_DATA WHERE PN = ?", (job_id_to_delete,))
                conn.commit()
                conn.close()

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'message': f'Job {job_id_to_delete} deleted successfully'}).encode('utf-8'))

            except ValueError as e:
                self.send_response(404)  # Not found for invalid job ID
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))   
                







        elif self.path == '/submit-job':
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                data = json.loads(post_data.decode('utf-8'))

                # Define the Excel sheet headers (same as your fieldnames)
                fieldnames = ['INPUT DATE', 'PN', 'NO/CELL', 'DRAW NO', 'REQU-DATE', 'CUST', 'STOCK CODE',
                            'QTY', 'CELL CODE', 'B$', 'ORDER NO', 'MODEL', 'VOL', 'AH', 'WH', 'CHEM',
                            'STRUCTURE', 'STAFF', 'WORKHR', 'HRPP', 'END DATE', 'TEST TIME', 'AV',
                            'S$', 'C-DRAW', 'C-CELLS', 'C-AV', 'C-B$', 'C-S$', 'C-STCODE', 'ORIGINAL', 'DISCOUNT',
                            'SALESMAN', 'CUSTOMERCODE', 'ORDERDATE']

                # Mapping the incoming JSON data to the Excel columns
                mapped_data = {
                    'INPUT DATE': data.get('inputDate', ''),
                    'PN': data.get('pn', ''),
                    'NO/CELL': data.get('noCell', ''),
                    'DRAW NO': data.get('drawNo', ''),
                    'REQU-DATE': data.get('requDate', ''),
                    'CUST': data.get('cust', ''),
                    'STOCK CODE': data.get('stockCode', ''),
                    'QTY': data.get('qty', ''),
                    'CELL CODE': data.get('cellCode', ''),
                    'B$': data.get('bPrice', ''),
                    'ORDER NO': data.get('orderNo', ''),
                    'MODEL': data.get('model', ''),
                    'VOL': data.get('vol', ''),
                    'AH': data.get('ah', ''),
                    'WH': data.get('wh', ''),
                    'CHEM': data.get('chem', ''),
                    'STRUCTURE': data.get('structure', ''),
                    'STAFF': data.get('staff', ''),
                    'WORKHR': data.get('workhr', ''),
                    'HRPP': data.get('HRPP', ''),
                    'END DATE': data.get('endDate', ''),
                    'TEST TIME': data.get('testTime', ''),
                    'AV': data.get('av', ''),
                    'S$': data.get('sPrice', ''),
                    'SALESMAN': data.get('salesman', ''),
                    'CUSTOMERCODE': data.get('customerCode', ''),
                    'ORDERDATE': data.get('orderDate', '')
                }

                # File path for the Excel file
                file_path = 'Sample.xlsx'

                # Check if the file exists
                file_exists = os.path.isfile(file_path)

                # Load or create the workbook
                if file_exists:
                    workbook = openpyxl.load_workbook(file_path)
                    sheet = workbook['Sheet1']
                else:
                    # Create a new workbook and write headers if file doesn't exist
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    # Write headers
                    sheet.append(fieldnames)

                # Append the mapped data to the sheet
                sheet.append([mapped_data.get(col, '') for col in fieldnames])

                # Save the workbook
                workbook.save(file_path)

                # Send success response
                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'message': 'Job added successfully!'}).encode('utf-8'))

            except Exception as e:
                # Handle errors
                self.send_response(500)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))

        elif self.path == '/finish-job':
            self.handle_finish_job()

        
        elif self.path == '/move-job':
            print("Handling /move-job request")  # Confirm route hit
            self.handle_move_job()


        #EDIT clock in/out
                
        elif self.path == '/edit-clock':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            data = json.loads(post_data.decode('utf-8'))
            
            # Admin check
            user_role = self.headers.get('X-User-Role', 'user').lower()
            if user_role != 'admin':
                self.send_response(403)
                self.set_cors_headers()  # ✅ Use helper function for consistent CORS
                self.end_headers()
                self.wfile.write(json.dumps({'error': 'Forbidden'}).encode())
                return

            record_id = data['recordId']
            new_start = data['newStartTime']
            new_stop = data['newStopTime']

            try:
                conn = get_db_connection()

                cursor = conn.cursor()

                # Check if the job is in progress (StopTime is NULL)
                cursor.execute("SELECT StopTime FROM ClockInOut WHERE RecordID = ?", (record_id,))
                stop_time = cursor.fetchone()

                if stop_time and stop_time[0] is None:
                    # The job is still running, only update StartTime
                    cursor.execute("""
                        UPDATE ClockInOut
                        SET StartTime = ?
                        WHERE RecordID = ? AND StopTime IS NULL
                    """, (new_start, record_id))
                else:
                    # If job is already stopped, allow full edit
                    cursor.execute("""
                        UPDATE ClockInOut
                        SET StartTime = ?, StopTime = ?
                        WHERE RecordID = ?
                    """, (new_start, new_stop, record_id))

                conn.commit()

                # --- Sync TestRecords if this job is finished ---
                cursor.execute("SELECT JobID FROM ClockInOut WHERE RecordID = ?", (record_id,))
                job_row = cursor.fetchone()
                if job_row:
                    update_testrecord_from_clock(job_row[0])

                self.send_response(200)
                self.set_cors_headers()  # ✅ Use helper function for consistent CORS
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'message': 'Record updated successfully'}).encode())
            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()  # ✅ Apply CORS headers for error response too
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode())
            finally:
                if conn:
                    conn.close()

        elif self.path == '/delete-clock':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            data = json.loads(post_data.decode('utf-8'))
            
            # Admin check
            user_role = self.headers.get('X-User-Role', 'user').lower()
            if user_role != 'admin':
                self.send_response(403)
                self.end_headers()
                return

            record_id = data['recordId']

            try:
                conn = get_db_connection()

                cursor = conn.cursor()
                # --- Fetch JobID before deleting, for TestRecords sync ---
                cursor.execute("SELECT JobID FROM ClockInOut WHERE RecordID = ?", (record_id,))
                job_row = cursor.fetchone()
                job_id = job_row[0] if job_row else None

                cursor.execute('DELETE FROM ClockInOut WHERE RecordID = ?', (record_id,))
                conn.commit()

                # --- Sync TestRecords if this job is finished ---
                if job_id:
                    update_testrecord_from_clock(job_id)

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'message': 'Record deleted successfully'}).encode())
            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode())
            finally:
                if conn:
                    conn.close()

       
        elif self.path == '/update-pn-row':
            conn = None
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                data = json.loads(post_data.decode('utf-8'))
                
                # Validate required fields
                if not data.get('pn') or not data.get('updates'):
                    self.send_response(400)
                    self.set_cors_headers()
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': "Both 'pn' and 'updates' are required"
                    }).encode())
                    return
                    
                conn = get_db_connection()

                cursor = conn.cursor()
                
                # Check if PN exists
                cursor.execute('SELECT PN FROM PN_DATA WHERE PN = ?', (data['pn'],))
                if not cursor.fetchone():
                    self.send_response(404)
                    self.set_cors_headers()
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': f"PN {data['pn']} not found"
                    }).encode())
                    return
                
                # Get all column names from the table
                # Get all column names from the table (case-sensitive)
                cursor.execute('PRAGMA table_info(PN_DATA)')
                all_columns = [col[1] for col in cursor.fetchall()]

                # Create a mapping of normalized column names to actual names
                column_mapping = {col.replace('_', '/').upper(): col for col in all_columns}

                # Validate updates
                set_clauses = []
                params = []

                for frontend_col, value in data['updates'].items():
                    # Normalize the frontend column name
                    normalized_col = frontend_col.replace('_', '/').upper()
                    
                    # Find matching database column
                    db_col = column_mapping.get(normalized_col)
                    if not db_col:
                        print(f"Ignoring unknown column: {frontend_col} (normalized: {normalized_col})")
                        continue
                        
                    # Numeric validation for QTY/AV
                    if db_col in ['QTY', 'AV']:
                        try:
                            float(value)
                        except ValueError:
                            self.send_response(400)
                            self.send_header('Content-Type', 'application/json')
                            self.end_headers()
                            self.wfile.write(json.dumps({
                                'success': False,
                                'error': f"{db_col} must be a number"
                            }).encode())
                            return
                            
                    set_clauses.append(f'"{db_col}" = ?')
                    params.append(value)

                
                if not set_clauses:
                    self.send_response(400)
                    self.set_cors_headers()
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': "No valid columns to update"
                    }).encode())
                    return
                
                params.append(data['pn'])
                
                # Execute update
                query = f'UPDATE PN_DATA SET {", ".join(set_clauses)} WHERE PN = ?'
                cursor.execute(query, params)
                conn.commit()
                
                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': True,
                    'message': 'Row updated successfully'
                }).encode())
                
            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': False,
                    'error': str(e)
                }).encode())
            finally:
                if conn:
                    conn.close()


        elif self.path == '/update-test-record':
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                data = json.loads(post_data.decode('utf-8'))

                pn = data.get('pn')
                raw_updates = data.get('updates', {})

                # Mapping for TestRecords
                column_map = {
                    'date': 'date',
                    'draw_no': 'draw_no',
                    'no/cell': 'cells',
                    'customer': 'customer',
                    'qty': 'qty',
                    'av': 'av',
                    'estimated_time': 'estimated_time',
                    'total_hours_worked': 'total_time',
                    'unit_price': 'unit_price',
                    'remaining_time': 'remaining_time',
                    'test_time': 'test_time',
                    'order_no': 'order_no',
                    'bill_price': 'bill_price',
                    'profit': 'profit',
                    'comment': 'comment',
                    'stock_code': 'stock_code',
                    'picture': 'picture',
                    'salesman': 'salesman',
                    'customer_code': 'customer_code'
                }

                # Mapping for JOBSFINISHED
                jobsfinished_column_map = {
                    'date': 'INPUT DATE',
                    'draw_no': 'DRAW NO',
                    'no/cell': 'NO/CELL',
                    'customer': 'CUST',
                    'qty': 'QTY',
                    'av': 'AV',
                    'estimated_time': 'EstimatedTime',
                    'total_hours_worked': 'TotalHoursWorked',
                    'unit_price': 'S$',
                    'remaining_time': 'RemainingTime',
                    'test_time': 'TEST TIME',
                    'order_no': 'ORDER NO',
                    'bill_price': 'B$',
                    'stock_code': 'STOCK CODE',
                    'salesman': 'SALESMAN',
                    'customer_code': 'Customer Code'
                }

                # Mapping for csv_data
                csv_column_map = {
                    'draw_no': 'Drawing_Number',
                    'qty': 'Qty',
                    'av': 'AVERAGE_TIME',
                    'comment': 'COMMENT',
                    'b_price': 'B_PRICE',
                    'bill_price': 'B_PRICE',       # ✅ alias for compatibility
                    's_price': 'S_PRICE',
                    'unit_price': 'S_PRICE'        # ✅ alias for compatibility
                }


                mapped_updates = {}
                job_updates = {}
                csv_updates = {}
                dynamic_fields = {}
                csv_draw_no = None

                for key, value in raw_updates.items():
                    if key in column_map:
                        mapped_updates[column_map[key]] = value
                    if key in jobsfinished_column_map:
                        job_updates[jobsfinished_column_map[key]] = value
                    if key in csv_column_map:
                        csv_col = csv_column_map[key]
                        csv_updates[csv_col] = value
                        if csv_col == 'Drawing_Number':
                            csv_draw_no = value
                    if key == 'draw_no':
                        csv_draw_no = value
                    if key.endswith('(hours_worked)') or key.endswith('(save_time)'):
                        dynamic_fields[key] = value

                if not mapped_updates and not dynamic_fields and not job_updates and not csv_updates:
                    raise ValueError("No valid fields to update.")

                conn = get_db_connection()

                cursor = conn.cursor()

                # ✅ Update TestRecords
                if mapped_updates:
                    set_clause = ", ".join([f'"{col}" = ?' for col in mapped_updates])
                    params = list(mapped_updates.values()) + [pn]
                    cursor.execute(f'UPDATE TestRecords SET {set_clause} WHERE PN = ?', params)

                # ✅ Update staff_details JSON in TestRecords
                if dynamic_fields:
                    cursor.execute('SELECT staff_details FROM TestRecords WHERE PN = ?', (pn,))
                    existing = cursor.fetchone()
                    staff_details = json.loads(existing[0]) if existing and existing[0] else {}
                    staff_details.update(dynamic_fields)
                    cursor.execute('UPDATE TestRecords SET staff_details = ? WHERE PN = ?', (json.dumps(staff_details), pn))

                    

                # ✅ Update JOBSFINISHED
                if job_updates:
                    job_set_clause = ", ".join([f'"{col}" = ?' for col in job_updates])
                    job_params = list(job_updates.values()) + [pn]
                    cursor.execute(f'UPDATE JOBSFINISHED SET {job_set_clause} WHERE PN = ?', job_params)
                    print("✅ JOBSFINISHED updated:", job_updates)

                # ✅ Update CSV_DATA using Drawing_Number and PN
                if csv_draw_no and pn and csv_updates:
                    print("🧩 Updating csv_data using DRAW_NO and PN")
                    print("  Drawing_Number =", csv_draw_no)
                    print("  PN =", pn)
                    print("  Updates =", csv_updates)

                    csv_set_clause = ", ".join([f'"{col}" = ?' for col in csv_updates])
                    csv_params = list(csv_updates.values()) + [csv_draw_no, pn]

                    cursor.execute(f'''
                        UPDATE csv_data SET {csv_set_clause}
                        WHERE Drawing_Number = ? AND PN = ?
                    ''', csv_params)
                    print("✅ CSV_DATA rows affected:", cursor.rowcount)
                
               # After updating TestRecords with mapped_updates...
                if 'total_time' in mapped_updates or 'total_hours_worked' in mapped_updates:
                    # Always use the field name in DB: total_time
                    cursor.execute('SELECT qty, unit_price, bill_price, total_time FROM TestRecords WHERE PN = ?', (pn,))
                    row = cursor.fetchone()
                    print(f"🟠 Calculating profit for PN={pn}: qty={row[0]}, unit_price={row[1]}, bill_price={row[2]}, total_time={row[3]}")
                    if row:
                        qty = float(row[0] or 0)
                        unit_price = float(row[1] or 0)
                        bill_price = float(row[2] or 0)
                        total_time = float(row[3] or 0)

                        # Fetch av for estimated_time calculation
                        cursor.execute('SELECT av FROM TestRecords WHERE PN = ?', (pn,))
                        av_row = cursor.fetchone()
                        av = float(av_row[0] or 0) if av_row else 0

                        estimated_time = qty * av
                        remaining_time = round(estimated_time - total_time, 2)
                        total_labor_cost = total_time * 37.95
                        profit = round((qty * unit_price) - total_labor_cost - (qty * bill_price), 2)
                        print(f"[DEBUG] Before profit update: PN={pn}, profit={profit}")
                        cursor.execute('UPDATE TestRecords SET profit = ? WHERE PN = ?', (profit, pn))
                        cursor.execute('UPDATE TestRecords SET remaining_time = ? WHERE PN = ?', (remaining_time, pn))
                        cursor.execute('SELECT profit FROM TestRecords WHERE PN = ?', (pn,))
                        print(f"[DEBUG] After profit update: PN={pn}, profit={cursor.fetchone()[0]}")
                        

                conn.commit()
                conn.close()

                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'message': 'Record updated successfully'}).encode('utf-8'))

            except Exception as e:
                print(f"❌ Error updating test record: {e}")
                self.send_response(500)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))

        elif self.path == '/update-multiple-test-records':
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                data = json.loads(post_data.decode('utf-8'))

                if not isinstance(data, list):
                    raise ValueError("Expected a list of update objects")

                # Mappings
                column_map = {
                    'date': 'date',
                    'draw_no': 'draw_no',
                    'no/cell': 'cells',
                    'customer': 'customer',
                    'qty': 'qty',
                    'av': 'av',
                    'estimated_time': 'estimated_time',
                    'total_hours_worked': 'total_time',
                    'unit_price': 'unit_price',
                    'remaining_time': 'remaining_time',
                    'test_time': 'test_time',
                    'order_no': 'order_no',
                    'bill_price': 'bill_price',
                    'profit': 'profit',
                    'comment': 'comment',
                    'stock_code': 'stock_code',
                    'picture': 'picture',
                    'salesman': 'salesman',
                    'customer_code': 'customer_code'
                }

                jobsfinished_column_map = {
                    'date': 'INPUT DATE',
                    'draw_no': 'DRAW NO',
                    'no/cell': 'NO/CELL',
                    'customer': 'CUST',
                    'qty': 'QTY',
                    'av': 'AV',
                    'estimated_time': 'EstimatedTime',
                    'total_hours_worked': 'TotalHoursWorked',
                    'unit_price': 'S$',
                    'remaining_time': 'RemainingTime',
                    'test_time': 'TEST TIME',
                    'order_no': 'ORDER NO',
                    'bill_price': 'B$',
                    'stock_code': 'STOCK CODE',
                    'salesman': 'SALESMAN',
                    'customer_code': 'Customer Code'
                }

                csv_column_map = {
                    'draw_no': 'Drawing_Number',
                    'qty': 'Qty',
                    'av': 'AVERAGE_TIME',
                    'comment': 'COMMENT',
                    'b_price': 'B_PRICE',
                    'bill_price': 'B_PRICE',
                    's_price': 'S_PRICE',
                    'unit_price': 'S_PRICE'
                }

                conn = get_db_connection()

                cursor = conn.cursor()

                for entry in data:
                    pn = entry.get("pn")
                    raw_updates = entry.get("updates", {})

                    if not pn or not isinstance(raw_updates, dict):
                        print(f"⚠️ Skipping invalid entry: {entry}")
                        continue

                    mapped_updates = {}
                    job_updates = {}
                    csv_updates = {}
                    dynamic_fields = {}
                    csv_draw_no = None

                    for key, value in raw_updates.items():
                        if key in column_map:
                            mapped_updates[column_map[key]] = value
                        if key in jobsfinished_column_map:
                            job_updates[jobsfinished_column_map[key]] = value
                        if key in csv_column_map:
                            csv_col = csv_column_map[key]
                            csv_updates[csv_col] = value
                            if csv_col == 'Drawing_Number':
                                csv_draw_no = value
                        if key == 'draw_no':
                            csv_draw_no = value
                        if key.endswith('(hours_worked)') or key.endswith('(save_time)'):
                            dynamic_fields[key] = value

                    # -- Update TestRecords --
                    if mapped_updates:
                        set_clause = ", ".join([f'"{col}" = ?' for col in mapped_updates])
                        params = list(mapped_updates.values()) + [pn]
                        cursor.execute(f'UPDATE TestRecords SET {set_clause} WHERE PN = ?', params)

                    # -- Update staff_details --
                    if dynamic_fields:
                        cursor.execute('SELECT staff_details FROM TestRecords WHERE PN = ?', (pn,))
                        existing = cursor.fetchone()
                        staff_details = {}
                        if existing and existing[0]:
                            try:
                                staff_details = json.loads(existing[0])
                            except Exception:
                                staff_details = {}

                        staff_details.update(dynamic_fields)
                        cursor.execute('UPDATE TestRecords SET staff_details = ? WHERE PN = ?', (json.dumps(staff_details), pn))

                    # -- Update JOBSFINISHED --
                    if job_updates:
                        set_clause = ", ".join([f'"{col}" = ?' for col in job_updates])
                        params = list(job_updates.values()) + [pn]
                        cursor.execute(f'UPDATE JOBSFINISHED SET {set_clause} WHERE PN = ?', params)

                    # -- Update csv_data --
                    if csv_draw_no and pn and csv_updates:
                        set_clause = ", ".join([f'"{col}" = ?' for col in csv_updates])
                        params = list(csv_updates.values()) + [csv_draw_no, pn]
                        cursor.execute(f'''
                            UPDATE csv_data SET {set_clause}
                            WHERE Drawing_Number = ? AND PN = ?
                        ''', params)

                        

                conn.commit()
                conn.close()


                self.send_response(200)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'message': 'All updates processed successfully'}).encode('utf-8'))

            except Exception as e:
                print(f"❌ Error in bulk update: {e}")
                self.send_response(500)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))


        elif self.path == '/set-backorder':
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                data = json.loads(post_data.decode('utf-8'))
                job_ids = data.get('jobIds', [])
                backorder = 1 if data.get('backorder', False) else 0

                if not job_ids:
                    self.send_response(400)
                    self.set_cors_headers()
                    self.end_headers()
                    self.wfile.write(json.dumps({'error': 'Missing job IDs'}).encode())
                    return

                conn = get_db_connection()

                cursor = conn.cursor()
                cursor.execute(
                    f'UPDATE PN_DATA SET backorder = ? WHERE PN IN ({",".join(["?"]*len(job_ids))})',
                    [backorder] + job_ids
                )
                conn.commit()
                conn.close()

                self.send_response(200)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'success': True}).encode())
            except Exception as e:
                self.send_response(500)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode())



        elif self.path == '/download-excel':
            self.handle_download_excel()


        elif self.path == '/download-pdf':
            self.handle_download_pdf()


        # In your do_POST method
        elif self.path == '/allocate-jobs':
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                data = json.loads(post_data.decode('utf-8'))

                job_ids = data.get('jobIds', [])
                staff_names = data.get('staffNames', [])

                if not job_ids or not staff_names:
                    self.send_response(400)
                    self.end_headers()
                    self.wfile.write(json.dumps({'error': 'Missing jobs or staff'}).encode())
                    return

                conn = get_db_connection()

                cursor = conn.cursor()

                # Get customer names for all jobs
                cursor.execute(f'''
                    SELECT PN, CUST FROM PN_DATA
                    WHERE PN IN ({','.join(['?']*len(job_ids))})
                ''', job_ids)
                job_customers = {row[0]: row[1] for row in cursor.fetchall()}

                # Insert all combinations
                for job_id in job_ids:
                    customer = job_customers.get(job_id, 'N/A')
                    for staff_name in staff_names:
                        cursor.execute('''
                            INSERT INTO AllocatedJobs 
                            (JobID, StaffName, AllocationDate, CustomerName)
                            VALUES (?, ?, ?, ?)
                        ''', (job_id, staff_name, get_current_timestamp(), customer))

                        # Fetch existing STAFF from PN_DATA
                        cursor.execute('SELECT STAFF FROM PN_DATA WHERE PN = ?', (job_id,))
                        existing_staff = cursor.fetchone()
                        existing_staff = existing_staff[0] if existing_staff and existing_staff[0] else ''

                        # Check if staff_name already exists (case insensitive)
                        existing_names = [name.strip().lower() for name in existing_staff.split(',') if name.strip()]
                        if staff_name.lower() not in existing_names:
                            if existing_staff:
                                updated_staff = f"{existing_staff}, {staff_name}"
                            else:
                                updated_staff = staff_name

                            cursor.execute('''
                                UPDATE PN_DATA
                                SET STAFF = ?
                                WHERE PN = ?
                            ''', (updated_staff, job_id))
                            print(f"✅ STAFF updated for PN {job_id}: {updated_staff}")
                        else:
                            print(f"⚠️ Staff '{staff_name}' already assigned to PN {job_id}, no update needed.")


                conn.commit()
                conn.close()

                self.send_response(200)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'success': True}).encode())

            except Exception as e:
                print(f"Error in bulk allocation: {e}")
                self.send_response(500)
                self.end_headers()
                self.wfile.write(json.dumps({'error': str(e)}).encode())

        

        elif self.path == '/upload-test-pdf':
            self.handle_upload_pdf()

    

    

       
        else:
            self.send_response(404)
            self.set_cors_headers()
            self.end_headers()




    
    

     #Finish job       
    def handle_finish_job(self):
        conn = None
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length)
            data = json.loads(post_data.decode('utf-8'))
            
            
            job_id = data.get('jobId')
            hourly_rate = 37.95  # Consider storing this in a config table

            if not job_id:
                self.send_response(400)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': 'Missing staffName or jobId'}).encode())
                return

            conn = get_db_connection()

            cursor = conn.cursor()

            # Ensure all entries for this job are closed
            cursor.execute('''
                UPDATE ClockInOut 
                SET StopTime = COALESCE(StopTime, ?)
                WHERE JobID = ? AND StopTime IS NULL
            ''', (get_current_timestamp(), job_id))

            # Calculate labor costs for all entries
            cursor.execute('''
                SELECT StaffName, StartTime, StopTime 
                FROM ClockInOut 
                WHERE JobID = ?
            ''', (job_id,))
            

            total_labor_cost = 0.0
            time_entries = cursor.fetchall()
            
            for staff_name ,start_time, stop_time in time_entries:
                try:
                    start_dt = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
                    stop_dt = datetime.strptime(stop_time, '%Y-%m-%d %H:%M:%S') if stop_time else datetime.now()
                    hours = (stop_dt - start_dt).total_seconds() / 3600
                    labor_cost = round(hours * hourly_rate, 2)
                    
                    
                    # Update individual record
                    cursor.execute('''
                        UPDATE ClockInOut
                        SET LaborCost = ?
                        WHERE StaffName = ? AND JobID = ? AND StartTime = ?
                    ''', (labor_cost, staff_name, job_id, start_time))
                    
                    total_labor_cost += labor_cost
                    
                except Exception as e:
                    print(f"Error processing entry {start_time}-{stop_time}: {str(e)}")
                    continue

            # Update job metadata
            estimated_time = round(calculate_estimated_time(job_id),2)
            total_hours_worked = round(get_total_hours_worked(job_id),2)
            remaining_time = round(estimated_time - total_hours_worked,2)
            total_labor_cost = round(total_labor_cost, 2)

            cursor.execute('''
                INSERT OR REPLACE INTO JobTable 
                (JobID, TotalLaborCost, EstimatedTime, TotalHoursWorked, RemainingTime,Status)
                VALUES (?, ?, ?, ?, ?,'Finished')
            ''', (job_id, total_labor_cost, estimated_time, total_hours_worked, remaining_time))

            conn.commit()
            
            self.send_response(200)
            self.set_cors_headers()
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({
                'message': 'Job finalized successfully',
                'totalLaborCost': total_labor_cost,
                'totalHours': total_hours_worked,
                'remainingTime': remaining_time
            }).encode())

        except sqlite3.Error as e:
            print(f"Database error: {str(e)}")
            if conn: conn.rollback()
            self.send_response(500)
            self.set_cors_headers()
            self.end_headers()
            self.wfile.write(json.dumps({'error': 'Database operation failed'}).encode())
            
        except Exception as e:
            print(f"Unexpected error: {str(e)}")
            if conn: conn.rollback()
            self.send_response(500)
            self.set_cors_headers()
            self.end_headers()
            self.wfile.write(json.dumps({'error': 'Internal server error'}).encode())
            
        finally:
            if conn: 
                conn.close()


    def handle_move_job(self):
        conn = None
        try:
            # Step 1: Parse request data
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length)
            data = json.loads(post_data.decode('utf-8'))
            job_id = data.get('jobId', '').strip()

            if not job_id:
                self.send_response(400)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': 'Job ID is required'}).encode())
                return

            print(f"Processing job_id: {job_id}")  # Debug

            hourly_rate = 37.95  # keep in sync with finish handler / config
            max_retries = 5
            retry_delay = 1

            for attempt in range(max_retries):
                try:
                    conn = sqlite3.connect(DB_NAME, timeout=30)
                    cursor = conn.cursor()

                    # --- Recompute hours fresh from ClockInOut to avoid any stale values ---
                    # Non-QC hours (used for TotalHoursWorked and Labor Cost)
                    cursor.execute('''
                        SELECT SUM((strftime('%s', COALESCE(StopTime, 'now')) - strftime('%s', StartTime)) / 3600.0)
                        FROM ClockInOut WHERE JobID = ? AND LOWER(StaffName) != 'qc'
                    ''', (job_id,))
                    non_qc_hours = cursor.fetchone()[0] or 0.0
                    non_qc_hours = round(non_qc_hours, 2)

                    # QC-only hours (TEST TIME)
                    cursor.execute('''
                        SELECT SUM((strftime('%s', COALESCE(StopTime, 'now')) - strftime('%s', StartTime)) / 3600.0)
                        FROM ClockInOut WHERE JobID = ? AND LOWER(StaffName) = 'qc'
                    ''', (job_id,))
                    test_time_qc = cursor.fetchone()[0] or 0.0
                    test_time_qc = round(test_time_qc, 2)

                    # Labor cost from non-QC hours only
                    total_labor_cost = round(non_qc_hours * hourly_rate, 2)

                    # Step 3: Check for special job and fetch qty + exclude flag + AV
                    cursor.execute('SELECT QTY, EXCLUDE_SAVE_TIME, AV FROM PN_DATA WHERE PN = ?', (job_id,))
                    row = cursor.fetchone()
                    qty = float(row[0]) if row and row[0] else 0.0
                    exclude = int(row[1]) if row and row[1] else 0
                    av_from_pn = float(row[2]) if row and row[2] else 0.0

                    # Determine AV to store in JOBSFINISHED
                    if exclude == 1:
                        av = round(non_qc_hours / qty, 5) if qty else 0.0
                        estimated_time = round(non_qc_hours, 2)
                        remaining_time = 0.0
                    else:
                        av = av_from_pn
                        estimated_time = round((qty * av), 2)
                        remaining_time = round(estimated_time - non_qc_hours, 2)

                    # Keep END DATE as today
                    end_date_today = datetime.now().strftime('%Y-%m-%d')

                    # --- (Optional but recommended) Update JobTable to the corrected values before copy ---
                    cursor.execute('''
                        INSERT OR REPLACE INTO JobTable
                        (JobID, TotalLaborCost, EstimatedTime, TotalHoursWorked, RemainingTime, Status)
                        VALUES (?, ?, ?, ?, ?, 'Finished')
                    ''', (job_id, total_labor_cost, estimated_time, non_qc_hours, remaining_time))

                    # Step 5: Insert into JOBSFINISHED with corrected values (non-QC cost/hours + QC test time)
                    cursor.execute('''
                        INSERT INTO JOBSFINISHED (
                            "INPUT DATE", PN, "NO/CELL", "DRAW NO", "REQU-DATE", CUST, "STOCK CODE", 
                            QTY, "CELL CODE", "B$", "ORDER NO", MODEL, VOL, AH, WH, CHEM, STRUCTURE, 
                            STAFF, WORKHR, "HR/PP", "END DATE", "TEST TIME", AV, "S$", "C-DRAW", 
                            "C-CELLS", "C-AV", "C-B$", "C-S$", "C-STCODE", "ORIGINAL S$", DISCOUNT, 
                            SALESMAN, "Customer Code", "Order Date", TotalLaborCost, EstimatedTime, 
                            TotalHoursWorked, RemainingTime, Status, EXCLUDE_SAVE_TIME, "PRODUCTION READY DATE"
                        )
                        SELECT 
                            "INPUT DATE", PN, "NO/CELL", "DRAW NO", "REQU-DATE", CUST, "STOCK CODE", 
                            QTY, "CELL CODE", "B$", "ORDER NO", MODEL, VOL, AH, WH, CHEM, STRUCTURE, 
                            STAFF, WORKHR, "HR/PP", ?, ?, ?, "S$", "C-DRAW", 
                            "C-CELLS", "C-AV", "C-B$", "C-S$", "C-STCODE", "ORIGINAL S$", DISCOUNT, 
                            SALESMAN, "Customer Code", "Order Date", ?, ?, ?, ?, 'Completed', EXCLUDE_SAVE_TIME, "PRODUCTION READY DATE"
                        FROM PN_DATA WHERE PN = ?
                    ''', (
                        end_date_today,           # "END DATE"
                        test_time_qc,             # "TEST TIME" (QC-only hours)
                        av,                       # AV to store
                        total_labor_cost,         # TotalLaborCost (non-QC only)
                        estimated_time,           # EstimatedTime
                        non_qc_hours,             # TotalHoursWorked (non-QC only)
                        remaining_time,           # RemainingTime
                        job_id
                    ))

                    # Move to TestRecords (uses values from JOBSFINISHED we just inserted)
                    cursor.execute('''
                        INSERT OR IGNORE INTO TestRecords (
                            date, PN, draw_no, cells, customer, qty, av, unit_price, bill_price,
                            estimated_time, total_time, remaining_time, test_time, stock_code,
                            total_labor_cost, salesman, customer_code, order_no, profit, staff_details
                        )
                        SELECT 
                            j."END DATE" as date,
                            j.PN,
                            j."DRAW NO" as draw_no,
                            CAST(j."NO/CELL" AS INTEGER) as cells,
                            j.CUST as customer,
                            CAST(j.QTY AS REAL) as qty,
                            CAST(j.AV AS REAL) as av,
                            REPLACE(j."S$", '$', '') as unit_price,
                            REPLACE(j."B$", '$', '') as bill_price,
                            j.EstimatedTime as estimated_time,  
                            j.TotalHoursWorked as total_time,
                            j.RemainingTime as remaining_time,
                            (
                                SELECT ROUND(SUM(
                                    (strftime('%s', COALESCE(StopTime, 'now')) - strftime('%s', StartTime)) / 3600.0
                                ), 2)
                                FROM ClockInOut
                                WHERE JobID = j.PN AND LOWER(StaffName) = 'qc'
                            ) as test_time,
                            j."STOCK CODE" as stock_code,
                            j.TotalLaborCost as total_labor_cost,
                            j.SALESMAN as salesman,
                            j."Customer Code" as customer_code,
                            j."ORDER NO" as order_no,
                            (CAST(REPLACE(j."S$", '$', '') AS REAL) * CAST(j.QTY AS REAL)
                            - CAST(REPLACE(j."B$", '$', '') AS REAL) * CAST(j.QTY AS REAL)
                            - j.TotalLaborCost) as profit,
                            '{}' -- Placeholder for staff_details
                        FROM JOBSFINISHED j WHERE j.PN = ?
                    ''', (job_id,))

                    # Step 6: Update csv_data and MergedData (unchanged, keep your logic here)

                    # Fetch PN_DATA details
                    cursor.execute('''
                        SELECT "DRAW NO", QTY, CUST, "B$", "S$", PN, "INPUT DATE", "NO/CELL", "STOCK CODE", "CELL CODE", MODEL
                        FROM PN_DATA WHERE PN = ?
                    ''', (job_id,))
                    pn_data = cursor.fetchone()

                    if pn_data:
                        drawing_number = pn_data[0] or 'N/A'
                        qty = int(pn_data[1]) if pn_data[1] else 0
                        cust = pn_data[2] or 'N/A'
                        b_price = float(pn_data[3]) if pn_data[3] else 0.0
                        s_price = float(pn_data[4]) if pn_data[4] else 0.0
                        pn = pn_data[5] or 'N/A'
                        input_date = pn_data[6] or '1900-01-01'
                        cells = int(pn_data[7]) if pn_data[7] and str(pn_data[7]).isdigit() else None
                        stock_code = pn_data[8] or 'N/A'
                        cells_parts = pn_data[9] or 'N/A'
                        model = pn_data[10] or 'N/A'

                        used_time = f"{non_qc_hours:.4f}"
                        current_av = round(non_qc_hours / qty, 5) if qty else 0.0

                        cursor.execute('SELECT DISTINCT StaffName FROM ClockInOut WHERE JobID = ? AND StaffName IS NOT NULL', (job_id,))
                        staff_names = [row[0] for row in cursor.fetchall()]
                        staff = "/".join(staff_names) if staff_names else "N/A"

                        cursor.execute('SELECT SUM(CAST(USED_TIME AS REAL)), SUM(Qty) FROM csv_data WHERE Drawing_Number = ?', (drawing_number,))
                        sum_used_time, sum_qty = cursor.fetchone() or (0.0, 0)
                        sum_used_time = float(sum_used_time or 0.0) + float(used_time)
                        sum_qty = int(sum_qty or 0) + qty
                        average = round(sum_used_time / sum_qty, 5) if sum_qty else 0.0

                        cursor.execute('''
                            INSERT INTO csv_data (
                                Drawing_Number, DATE, Qty, USED_TIME, CURRENT_AV,
                                AVERAGE_TIME, STAFF, COMMENT, NEW, 'TOTAL_AV',
                                CUST, CELLS, B_PRICE, S_PRICE, PN
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            drawing_number, input_date, qty, used_time, current_av,
                            average, staff, 'N/A', 'N/A', average,
                            cust, cells, b_price, s_price, pn
                        ))

                        cursor.execute('''
                            INSERT INTO MergedData (DrawNo, AV, STOCKCODE, CELLS_PARTS, MODEL)
                            VALUES (?, ?, ?, ?, ?)
                            ON CONFLICT(DrawNo) DO UPDATE SET 
                                AV = excluded.AV,
                                STOCKCODE = excluded.STOCKCODE,
                                CELLS_PARTS = excluded.CELLS_PARTS,
                                MODEL = excluded.MODEL
                        ''', (drawing_number, average, stock_code, cells_parts, model))

                    # Step 7: Cleanup
                    cursor.execute('DELETE FROM PN_DATA WHERE PN = ?', (job_id,))
                    cursor.execute('DELETE FROM JobTable WHERE JobID = ?', (job_id,))
                    cursor.execute('DELETE FROM AllocatedJobs WHERE JobID = ?', (job_id,))
                    conn.commit()
                    break

                except sqlite3.OperationalError as e:
                    if "database is locked" in str(e).lower():
                        print(f"Database locked, retrying ({attempt + 1}/{max_retries})...")
                        time.sleep(retry_delay)
                        continue
                    raise
                except Exception as e:
                    print(f"Error during move-job: {str(e)}")
                    if conn:
                        conn.rollback()
                    raise

            self.send_response(200)
            self.set_cors_headers()
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'message': 'Job moved successfully!'}).encode())

        except Exception as e:
            print(f"Unexpected error: {str(e)}")
            self.send_response(500)
            self.set_cors_headers()
            self.end_headers()
            self.wfile.write(json.dumps({'error': 'Internal server error'}).encode())

        finally:
            if conn:
                conn.close()


    

    def get_av_value(self, draw_no):
        """Search for the DRAW NO and return the corresponding AV value from the Excel file."""
        try:
            # Load the Excel file and open the sheet
            wb = openpyxl.load_workbook('SampleAV.xlsx')
            sheet = wb['MergedData']  
            
            # Iterate through rows to find matching DRAW NO
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
                if row[0].value == draw_no:
                    return row[1].value  # Assuming AV value is in the second column
            return None  # If DRAW NO not found
        except Exception as e:
            print(f"Error reading Excel file: {str(e)}")
            return None

    
    # PRODUCTION TEST RECORDS
    def handle_get_test_records(self):
        try:
            print("Fetching test records...")
            self.send_response(200)
            self.set_cors_headers()
            self.send_header('Content-type', 'application/json')
            self.end_headers()

            conn = get_db_connection()

            cursor = conn.cursor()

            # Create the TestRecords table if it doesn't exist
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS TestRecords (
                    date TEXT,
                    PN TEXT PRIMARY KEY,
                    draw_no TEXT,
                    cells INTEGER,
                    customer TEXT,
                    qty REAL,
                    av REAL,
                    unit_price REAL,
                    bill_price REAL,
                    estimated_time REAL,
                    total_time REAL,
                    remaining_time REAL,
                    test_time REAL,
                    stock_code TEXT,
                    picture TEXT,  
                    total_labor_cost REAL,
                    salesman TEXT,
                    customer_code TEXT,
                    order_no TEXT,
                    profit REAL,
                    comment TEXT DEFAULT '',
                    staff_details TEXT,
                    has_pdf INTEGER DEFAULT 0
                )
            ''')

            # Insert data from JOBSFINISHED into TestRecords
            cursor.execute('''
            INSERT OR IGNORE INTO TestRecords (
                date, PN, draw_no, cells, customer, qty, av, unit_price, bill_price,
                estimated_time, total_time, remaining_time, test_time, stock_code,
                total_labor_cost, salesman, customer_code, order_no, profit, staff_details
            )
            SELECT 
                j."END DATE" as date,
                j.PN,
                j."DRAW NO" as draw_no,
                CAST(j."NO/CELL" AS INTEGER) as cells,
                j.CUST as customer,
                CAST(j.QTY AS REAL) as qty,
                CAST(j.AV AS REAL) as av,
                REPLACE(j."S$", '$', '') as unit_price,  -- Remove dollar sign
                REPLACE(j."B$", '$', '') as bill_price,  -- Remove dollar sign
                j.EstimatedTime as estimated_time,  
                j.TotalHoursWorked as total_time,
                j.RemainingTime as remaining_time,
                (
                    SELECT ROUND(SUM(
                        (strftime('%s', COALESCE(StopTime, 'now')) - strftime('%s', StartTime)) / 3600.0
                    ), 2)
                    FROM ClockInOut
                    WHERE JobID = j.PN AND StaffName = 'QC'
                ) as test_time,
                j."STOCK CODE" as stock_code,
                j.TotalLaborCost as total_labor_cost,
                j.SALESMAN as salesman,
                j."Customer Code" as customer_code,
                j."ORDER NO" as order_no,
                (CAST(REPLACE(j."S$", '$', '') AS REAL) * CAST(j.QTY AS REAL) - CAST(REPLACE(j."B$", '$', '') AS REAL) * CAST(j.QTY AS REAL) - j.TotalLaborCost) as profit,
                '{}' -- Placeholder for staff_details
            FROM JOBSFINISHED j;
            ''')
            conn.commit()

            # Fetch records from the TestRecords table
            cursor.execute('SELECT * FROM TestRecords')
            columns = [col[0] for col in cursor.description]
            records = []

            for row in cursor.fetchall():
                record = dict(zip(columns, row))
                
                # Safely convert and assign values, defaulting when needed
                record['qty'] = float(record.get('qty') or 0)
                record['av'] = float(record.get('av') or 0)
                record['unit_price'] = float(record.get('unit_price') or 0)
                record['bill_price'] = float(record.get('bill_price') or 0)
                record['estimated_time'] = round(float(record.get('estimated_time') or 0), 2)
                record['total_time'] = round(float(record.get('total_time') or 0), 2)
                record['remaining_time'] = round(float(record.get('remaining_time') or 0), 2)
                record['test_time'] = round(float(record.get('test_time') or 0), 2)
                record['total_labor_cost'] = round(float(record.get('total_labor_cost') or 0), 2)

                record['profit'] = round(float(record.get('profit') or 0), 2)

                # Ensure optional fields have fallback values
                record['draw_no'] = record.get('draw_no', '')
                record['cells'] = record.get('cells', 0)
                record['customer'] = record.get('customer', '')
                record['stock_code'] = record.get('stock_code', '')
                record['picture'] = record.get('picture', '')
                record['salesman'] = record.get('salesman', '')
                record['customer_code'] = record.get('customer_code', '')
                record['order_no'] = record.get('order_no', '')
                record['date'] = record.get('date', '')
                record['has_pdf'] = 1 if record.get('picture') else 0

                if record['picture']:
                    record['pdf_name'] = os.path.basename(record['picture'])
                else:
                    record['pdf_name'] = None

                # Fetch staff work details for the current job
                cursor.execute('''
                SELECT StaffName, 
                    ROUND(SUM((strftime('%s', COALESCE(StopTime, 'now')) - strftime('%s', StartTime)) / 3600.0), 2) as worked_hours
                FROM ClockInOut
                WHERE JobID = ?
                GROUP BY StaffName
                ''', (record['PN'],))

                staff_work_details = cursor.fetchall()
                staff_details = {staff_name: {'worked_hours': worked_hours} for staff_name, worked_hours in staff_work_details}
                record['staff_details'] = json.dumps(staff_details)

                records.append(record)
            records.sort(key=lambda x: x['date'] or '', reverse=True)

            # Send the JSON response
            self.wfile.write(json.dumps(records).encode('utf-8'))
            print("Test records fetched and stored successfully.")

        except sqlite3.Error as e:
            print(f"Database error in handle_get_test_records: {e}")
            self.send_response(500)
            self.set_cors_headers()
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'error': 'Database error', 'details': str(e)}).encode('utf-8'))

        except Exception as e:
            print(f"Unexpected error in handle_get_test_records: {e}")
            self.send_response(500)
            self.set_cors_headers()
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'error': 'Unexpected error', 'details': str(e)}).encode('utf-8'))

        finally:
            if conn:
                conn.close()




    def handle_update_test_time(self):
        try:
            content_length = int(self.headers['Content-Length'])
            post_data = json.loads(self.rfile.read(content_length))
            
            conn = get_db_connection()

            cursor = conn.cursor()
            
            cursor.execute('''
            INSERT OR REPLACE INTO PROD_TEST_RECORDS 
            (pn, test_time)
            VALUES (?, ?)
            ''', (post_data['pn'], post_data['test_time']))
            
            conn.commit()
            self.send_response(200)
            self.set_cors_headers()
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())
            
        except Exception as e:
            self.send_error(500, str(e))
        finally:
            conn.close()


    # Upload PDF in products test record.
    

    def handle_upload_pdf(self):
        try:
            content_type = self.headers.get('Content-Type', '')
            if not content_type.startswith('multipart/form-data'):
                self.send_response(400)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': 'Invalid content type'}).encode())
                return

            # Read the content length
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                self.send_response(400)
                self.set_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'error': 'Empty request body'}).encode())
                return

            # Read the raw body
            body = self.rfile.read(content_length)

            # Parse the multipart data manually
            boundary = content_type.split('boundary=')[1].encode()
            parts = body.split(b'--' + boundary)

            pn = None
            file_data = None
            filename = None

            for part in parts:
                if b'Content-Disposition' in part:
                    headers, data = part.split(b'\r\n\r\n', 1)
                    headers = headers.decode('utf-8')
                    data = data.rstrip(b'\r\n--')

                    if 'name="pn"' in headers:
                        pn = data.decode('utf-8').strip()
                    elif 'name="file"' in headers and 'filename=' in headers:
                        filename = headers.split('filename="')[1].split('"')[0]
                        file_data = data

            if not pn or not file_data:
                self.send_response(400)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': 'PN and file are required'}).encode())
                return

            # Validate file type
            if not filename.lower().endswith('.pdf'):
                self.send_response(400)
                self.set_cors_headers()
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': 'Only PDF files are allowed'}).encode())
                return

            # Save the file
            file_path = os.path.join(PRODUCT_PHOTOS, f"{pn}.pdf")
            with open(file_path, 'wb') as f:
                f.write(file_data)

                    # Update TestRecords directly
            conn = get_db_connection()

            cursor = conn.cursor()
            cursor.execute('''
                UPDATE TestRecords
                SET picture = ?, has_pdf =1
                WHERE PN = ?
            ''', (file_path, pn))
            conn.commit()
            conn.close()


            # Respond with success
            self.send_response(200)
            self.set_cors_headers()
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'message': 'PDF uploaded successfully'}).encode())

        except Exception as e:
            print(f"Error in handle_upload_pdf: {str(e)}")
            self.send_response(500)
            self.set_cors_headers()
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode())




    def handle_update_test_time(self):
        try:
            content_length = int(self.headers['Content-Length'])
            post_data = json.loads(self.rfile.read(content_length))
            
            conn = get_db_connection()

            cursor = conn.cursor()
            
            # Update only the columns that exist in the schema
            set_clause = ", ".join([f'"{col}" = ?' for col in [
                'date', 'PN', 'draw_no', 'cells', 'customer', 
                'qty', 'av', 'remaining_time', 'test_time',
                'order_no', 'stock_code', 'salesman', 'customer_code'
            ]])
            
            values = [
                post_data.get('DATE', ''),
                post_data.get('PN', ''),
                post_data.get('DRAW NO', ''),
                post_data.get('NO/CELL', ''),
                post_data.get('CUSTOMER', ''),
                post_data.get('QTY', '0'),
                post_data.get('AV', '0'),
                post_data.get('REMAINING TIME', '0'),
                post_data.get('TEST TIME', ''),
                post_data.get('ORDER NO', ''),
                post_data.get('STOCK CODE', ''),
                post_data.get('SALESMAN', ''),
                post_data.get('CUSTOMER CODE', ''),
                post_data.get('PN')  # PN is the WHERE clause parameter
            ]
            
            query = f'''
                UPDATE TestRecords
                SET {set_clause}
                WHERE PN = ?
            '''
            
            cursor.execute(query, values)
            conn.commit()
            
            self.send_response(200)
            self.set_cors_headers()
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'message': 'Record updated'}).encode())

        except Exception as e:
            print(f"❌ Error updating test record: {e}")  # FULL ERROR LOG
            self.send_response(500)
            self.set_cors_headers()
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode())
        finally:
            if conn:
                conn.close()


    def handle_download_excel(self):
        try:
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            data = json.loads(post_data.decode('utf-8'))
            
            # Connect to database
            conn = get_db_connection()

            cursor = conn.cursor()
            
            # Base query
            query = '''
                SELECT * FROM TestRecords
                WHERE 1=1
            '''
            params = []
            
            # Add date filter if provided
            if data.get('start_date') and data.get('end_date'):
                query += ' AND date BETWEEN ? AND ?'
                params.extend([data['start_date'], data['end_date']])
            
            # Add staff filter if provided
            if data.get('staff_name'):
                query += ' AND staff_details LIKE ?'
                params.append(f'%{data["staff_name"]}%')
            
            # Execute query
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            # Get column names
            columns = [col[0] for col in cursor.description]
            
            # Create DataFrame
            df = pd.DataFrame(rows, columns=columns)
            
            # Create Excel file in memory
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Production Records')
            
            # Prepare response
            self.send_response(200)
            self.set_cors_headers()
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="Production_Test_Records.xlsx"')
            self.end_headers()
            self.wfile.write(output.getvalue())
            
        except Exception as e:
            self.send_response(500)
            self.set_cors_headers()
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode())


    
    def handle_download_pdf(self):
        try:
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            data = json.loads(post_data.decode('utf-8'))
            
            # Connect to database (same filtering as Excel)
            conn = get_db_connection()

            cursor = conn.cursor()
            
            query = '''
                SELECT * FROM TestRecords
                WHERE 1=1
            '''
            params = []
            
            if data.get('start_date') and data.get('end_date'):
                query += ' AND date BETWEEN ? AND ?'
                params.extend([data['start_date'], data['end_date']])
            
            if data.get('staff_name'):
                query += ' AND staff_details LIKE ?'
                params.append(f'%{data["staff_name"]}%')
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            
            # Create PDF
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
            from reportlab.lib import colors
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            
            # Prepare data for table
            table_data = [columns]  # headers
            table_data.extend(rows)  # data
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            # Build PDF
            doc.build([table])
            
            # Prepare response
            self.send_response(200)
            self.set_cors_headers()
            self.send_header('Content-Type', 'application/pdf')
            self.send_header('Content-Disposition', 'attachment; filename="Production_Test_Records.pdf"')
            self.end_headers()
            self.wfile.write(buffer.getvalue())


            
        except Exception as e:
            self.send_response(500)
            self.set_cors_headers()
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode())







# Run the server
def run(server_class=HTTPServer, handler_class=ClockInOutHandler, port=PORT):
    server_address = ('0.0.0.0', port)  # Listen on all interfaces
    httpd = server_class(server_address, handler_class)
    httpd = server_class(server_address, handler_class)
    print(f"🚀 Environment: {ENV}")
    print(f"🌐 Running on port: {port}")
    print(f"✅ Server started on port {port}...")
    httpd.serve_forever()

if __name__ == '__main__':
    run()
