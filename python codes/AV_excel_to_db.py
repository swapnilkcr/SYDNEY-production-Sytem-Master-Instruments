import sqlite3
from openpyxl import load_workbook

# Load Excel file
wb = load_workbook(r'S:\DATAPERTH\Perth Log Job system\Clock_In_test\app\SampleAV_Updated.xlsx')
sheet = wb['MergedData']

# Connect to database
conn = sqlite3.connect('clock_in_management.db')
cursor = conn.cursor()

# Skip header row, start from second
for row in sheet.iter_rows(min_row=2, values_only=True):
    draw_no, av, stockcode, cells_parts, model = row

    # Optional: skip empty draw values
    if not draw_no:
        continue

    cursor.execute('''
    INSERT INTO MergedData (DrawNo, AV, STOCKCODE, CELLS_PARTS, MODEL)
    VALUES (?, ?, ?, ?, ?)
    ON CONFLICT(DrawNo) DO UPDATE SET
        AV = excluded.AV,
        STOCKCODE = excluded.STOCKCODE,
        CELLS_PARTS = excluded.CELLS_PARTS,
        MODEL = excluded.MODEL
''', (draw_no, av, stockcode, cells_parts, model))

# Commit and close
conn.commit()
conn.close()
print("✅ All rows from MergedData inserted successfully.")
